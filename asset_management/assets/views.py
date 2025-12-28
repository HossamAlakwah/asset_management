import io
from datetime import date
from itertools import count

#from turtle import Screen
from urllib.parse import unquote, urlparse

import pandas as pd
import xlsxwriter
from django.apps import apps
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.forms import inlineformset_factory
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .download_templates import (
    generate_access_point_template,
    generate_asset_template,
    generate_camera_template,
    generate_firewall_template,
    generate_nvr_template,
    generate_raya_datacenter_template,
    generate_router_template,
    generate_screen_template,
    generate_switch_template,
    generate_telephone_template,
    generate_ups_template,
    generate_zk_template,
)
from .extract_date import (
    generate_access_points_report,
    generate_assets_report,
    generate_cameras_report,
    generate_firewalls_report,
    generate_nvrs_report,
    generate_raya_report,
    generate_routers_report,
    generate_screens_report,
    generate_switches_report,
    generate_telephones_report,
    generate_ups_report,
    generate_zk_report,
)
from .forms import (  # ✅ assuming you already created these
    AccessPointEditForm,
    AccessPointForm,
    AssetForm,
    CameraEditForm,
    CameraForm,
    FirewallEditForm,
    FirewallForm,
    NotificationConfigForm,
    NotificationRecipientForm,
    NVREditForm,
    NVRForm,
    RayaDataCenterVMForm,
    RouterEditForm,
    RouterForm,
    ScreenForm,
    ServerEditForm,
    ServerForm,
    StorageDeviceFormSet,
    SwitchEditForm,
    SwitchForm,
    TelephoneForm,
    UPSEditForm,
    UPSForm,
    VirtualMachineEditForm,
    VirtualMachineForm,
    ZKDeviceEditForm,
    ZKDeviceForm,
)
from .models import (  # Adjust as needed
    NVR,
    UPS,
    AccessPoint,
    AccessPointLog,
    Asset,
    AssetLog,
    Branch,
    Camera,
    CameraLog,
    Employee,
    Firewall,
    FirewallLog,
    NotificationConfig,
    NotificationRecipient,
    NVRLog,
    RayaDataCenterVM,
    ReportableField,
    ReportableModel,
    Router,
    RouterLog,
    Screen,
    ScreenLog,
    Server,
    ServerLog,
    StorageDevice,
    Switch,
    SwitchLog,
    Telephone,
    TelephoneLog,
    UPSLog,
    VirtualMachine,
    VirtualMachineLog,
    ZKDevice,
    ZKDeviceLog,
)
from .upload_data import (
    upload_bulk_access_points,
    upload_bulk_asset,
    upload_bulk_cameras,
    upload_bulk_employee,
    upload_bulk_firewalls,
    upload_bulk_nvrs,
    upload_bulk_raya_vms,
    upload_bulk_routers,
    upload_bulk_screens,
    upload_bulk_switches,
    upload_bulk_telephones,
    upload_bulk_ups,
    upload_bulk_zk_devices,
)


def logout_view(request):
    # Log out the user
    logout(request)

    # Clear session
    request.session.flush()

    # Prepare response with redirect
    response = redirect('/')

    # Optional: delete sessionid and csrftoken cookies
    response.delete_cookie('sessionid')
    response.delete_cookie('csrftoken')

    return response

@login_required
def branches_view(request):
    branches = Branch.objects.all().order_by('id')
    return render(request, 'branches.html', {'branches': branches})

@login_required
def view_branch(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    return render(request, 'branch_detail.html', {'branch': branch})

@login_required
def view_all(request):
    return render(request, 'all.html')


''' 

employee view and create employee
This view allows users to create a new employee and assign an asset if available.

'''
@login_required
def employees_view(request):
    unique_departments = list(
        Employee.objects.values_list('department', flat=True).distinct()
    )
    print(unique_departments)
    print('---------')
    if request.method == 'POST':
        name = request.POST.get('name')
        department = request.POST.get('department')
        title = request.POST.get('title')
        email = request.POST.get('email')
        asset_id = request.POST.get('assigned_asset')

        employee = Employee.objects.create(
            name=name,
            department=department,
            title=title,
            email=email,
            created_by=request.user
        )

        if asset_id:
            try:
                asset = Asset.objects.get(pk=asset_id)
                asset.employee_name = employee
                asset.status = 'In Use'
                asset.save()
            except Asset.DoesNotExist:
                pass

        return redirect('employees')

    employees = Employee.objects.all()
    assets = Asset.objects.filter(status='Stock', employee_name__isnull=True)
    return render(request, 'employees/employees.html', {
        'employees': employees,
        'assets': assets,
        'unique_departments':unique_departments
    })

@login_required
def create_employee(request):
    assets = Asset.objects.filter(status='Stock')  # Only unassigned assets
    branches = Branch.objects.filter(choosable=True)
    if request.method == 'POST':
        name = request.POST.get('name')
        department = request.POST.get('department')
        title = request.POST.get('title')
        email = request.POST.get('email')
        branch=request.POST.get('branch')
        branch = Branch.objects.filter(id=branch).first() 
        assigned_asset_id = request.POST.get('assigned_asset')

        if not all([name, department, title, email]):
            messages.error(request, "All fields are required.")
            return redirect('create_employee')

        try:
            with transaction.atomic():
                employee = Employee.objects.create(
                    name=name,
                    department=department,
                    title=title,
                    email=email,
                    branch=branch,
                    created_by=request.user
                )

            if assigned_asset_id:
                # Check if already assigned (avoid duplicates)
                already_assigned = employee.asset_set.filter(id=assigned_asset_id).exists()
                if not already_assigned:
                    asset = Asset.objects.get(pk=assigned_asset_id)
                    asset.employee_name = employee
                    asset.status = 'In Use'
                    asset.branch = employee.branch
                    asset.on_hand_date = timezone.now().date()
                    asset._changed_by = request.user
                    asset.save()


                messages.success(request, f"Employee '{employee.name}' created successfully.")
                return redirect('all_assets')

        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, 'employees/create_employee.html', {
        'assets': assets,
        'branches': branches,  # Pass branches to the template
    })




@login_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    assets = Asset.objects.filter(status='Stock').exclude(employee_name__isnull=False)
    screens = Screen.objects.filter(status='Stock').exclude(employee__isnull=False)
    telephones = Telephone.objects.filter(status='Stock').exclude(employee__isnull=False)
    branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        name = request.POST.get('name')
        department = request.POST.get('department')
        title = request.POST.get('title')
        email = request.POST.get('email')
        branch_id = request.POST.get('branch')
        assigned_asset_id = request.POST.get('assigned_asset')
        assigned_screen_id = request.POST.get('assigned_screen')  # New
        assigned_telephone_id = request.POST.get('assigned_telephone')

        if not all([name, department, title, email]):
            messages.error(request, "All fields are required.")
            return redirect('edit_employee', employee_id=employee.id)

        try:
            with transaction.atomic():
                employee.name = name
                employee.department = department
                employee.title = title
                employee.email = email
                employee.branch = Branch.objects.filter(id=branch_id).first()
                employee.save()

                # Assign Asset if selected
                if assigned_asset_id:
                    already_assigned = employee.asset_set.filter(id=assigned_asset_id).exists()
                    if not already_assigned:
                        asset = Asset.objects.get(pk=assigned_asset_id)
                        asset.employee_name = employee
                        asset.status = 'In Use'
                        asset.branch = employee.branch
                        asset.on_hand_date = timezone.now().date()
                        asset._changed_by = request.user
                        asset.save()

                # Assign Screen if selected
                if assigned_screen_id:
                    screen = Screen.objects.get(pk=assigned_screen_id)
                    screen.employee = employee
                    screen.status = 'In Use'
                    screen.branch = employee.branch
                    screen._changed_by = request.user
                    screen.save()

                if assigned_telephone_id:
                    telephone = Telephone.objects.get(pk=assigned_telephone_id)
                    telephone.employee = employee
                    telephone.status = 'In Use'
                    telephone.branch = employee.branch
                    telephone._changed_by = request.user
                    telephone.updated_at = timezone.now()
                    telephone.save()
                messages.success(request, f"Employee '{employee.name}' updated successfully.")
                return redirect('edit_employee', employee_id=employee.id)

        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, 'employees/edit_employee.html', {
        'employee': employee,
        'assets': assets,
        'screens': screens,
        'telephones':telephones,
        'branches': branches,
        # 'assigned_screens': assigned_screens,
    })




@require_POST
@login_required
def unassign_asset(request, asset_id, employee_id):
    asset = get_object_or_404(Asset, id=asset_id, employee_name__id=employee_id)
    stock_branch = Branch.objects.get(name__iexact='Stock')

    try:
        asset.employee_name = None
        asset.status = 'Stock'
        asset.branch = stock_branch
        asset.return_date = timezone.now().date()
        asset._changed_by = request.user
        asset.save()

        messages.success(request, f"Asset '{asset.serial}' unassigned successfully.")
    except Exception as e:
        messages.error(request, f"Error while unassigning asset: {e}")

    return redirect('edit_employee', employee_id=employee_id)

@login_required
def all_assets(request):
    assets = Asset.objects.all()
    all_count = assets.count()
    laptops_total = assets.filter(type='Laptop').count()
    PCs_total = assets.filter(type='Desktop').count()
    unique_asset_types = list(
        assets.order_by('status').values_list('status', flat=True).distinct()
    )
    print(unique_asset_types)
    stock_count = assets.filter(status='Stock').count()
    in_use_count = assets.filter(status='In Use').count()
    damage_count = assets.filter(status='Damage').count()
    
    context = {
        'branch': None,
        'assets': assets,
        'all': all_count,
        'laptops_total': laptops_total,
        'PCs_total': PCs_total,
        "unique_asset_types": unique_asset_types,
        "stock_count": stock_count,
        "in_use_count": in_use_count,
        "damage_count": damage_count,
    }
    return render(request, 'assets/assets_all.html', context)



@login_required
def download_asset_template(request):
    return generate_asset_template()

@login_required
def all_assets_log(request, slug):
    if slug == 'All':
        logs = AssetLog.objects.all().order_by('-change_time')
        context = {
            'logs': logs,
            'branch': None,
            'current_branch_slug': 'All',
        }
    else:
        branch = get_object_or_404(Branch, slug=slug)
        print(branch)
        asset_ids = Asset.objects.filter(branch=branch).values_list('id', flat=True)
        logs = AssetLog.objects.filter(asset_id__in=asset_ids).order_by('-change_time')

        context = {
            'logs': logs,
            'branch': branch,
            'current_branch_slug': branch.slug,
        }

    return render(request, 'assets/assets_logs.html', context)


@login_required
def branch_assets(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    assets = Asset.objects.filter(branch=branch)
    all_assets_count = assets.count()
    laptops_total = assets.filter(type='Laptop').count()
    PCs_total = assets.filter(type='Desktop').count()

    # Get unique asset types for the dropdown filter
    unique_asset_types = list(
        assets.order_by('type').values_list('type', flat=True).distinct()
    )

    logs = []

    context = {
        'assets': assets,
        'all': all_assets_count,
        'laptops_total': laptops_total,
        'PCs_total': PCs_total,
        'unique_asset_types': unique_asset_types,
        'logs': logs,
        'current_branch_slug': slug, 
    }
    return render(request, 'assets/branch_assets.html', context)

'''

upload bulk asset data from excel file 
laptops and PCs

'''

@require_POST 
@login_required
def upload_assets(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']

        success = upload_bulk_asset(request, excel_file, branch, slug, request.user)

        if not success:
            if slug!='All':
                return redirect('branch_assets', slug=slug)
            else:
                return redirect('all_assets')

    else:
        messages.error(request, "No Excel file was uploaded.")

    if slug!='All':
        return redirect('branch_assets', slug=slug)
    else:
        return redirect('all_assets')

@login_required
def extract_assets_data(request, slug):

    # Handle the special "All" case
    if slug == "All":
        branch = "All"
        print('all')
    else:
        branch = get_object_or_404(Branch, slug=slug)
        
    selected_status = request.POST.get('status') 
    selected_format = request.POST.get('format') 

    # Call the helper function to generate the report
    response = generate_assets_report(request, branch, selected_status, selected_format)

    if not response:
        return redirect('branch_assets', slug=slug)

    return response

@xframe_options_exempt
@login_required
def asset_details(request, asset_id):
    asset = get_object_or_404(Asset, pk=asset_id)
    print(asset)
    return render(request, 'assets/asset_details.html', {'asset': asset})



@login_required
def laptop_acknowledgment(request):
    return render(request, 'assets/test.html')




@login_required
def edit_asset(request, asset_id):
    asset = get_object_or_404(Asset, pk=asset_id)
    branches = Branch.objects.filter(choosable=True)
    employees = Employee.objects.all()

    if request.method == 'POST':
        # Fetch form fields
        product = request.POST.get('product')
        serial = request.POST.get('serial')
        status = request.POST.get('status')
        asset_type = request.POST.get('type')
        comments = request.POST.get('comments')
        branch_id = request.POST.get('branch')
        employee_id = request.POST.get('employee_name')

        warranty_str = request.POST.get('warranty')
        on_hand_date_str = request.POST.get('on_hand_date')
        return_date_str = request.POST.get('return_date')
        cpu = request.POST.get('cpu')
        cpu_gen = request.POST.get('cpu_generation')
        ram = request.POST.get('ram')

        # Validation
        missing_fields = []
        if not product: missing_fields.append("Product")
        if not serial: missing_fields.append("Serial")
        if not status: missing_fields.append("Status")
        if not asset_type: missing_fields.append("Asset Type")

        if missing_fields:
            messages.error(request, f"Please fill in: {', '.join(missing_fields)}")
            return render(request, 'assets/asset_edit.html', {
                'asset': asset,
                'branches': branches,
                'employees': employees,
                'cpu_choices': Asset.CPU_CHOICES,
                'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
                'ram_choices': Asset.RAM_CHOICES,
            })

        # Parse dates
        try:
            warranty = timezone.datetime.strptime(warranty_str, '%Y-%m-%d').date() if warranty_str else None
            on_hand_date = timezone.datetime.strptime(on_hand_date_str, '%Y-%m-%d').date() if on_hand_date_str else None
            return_date = timezone.datetime.strptime(return_date_str, '%Y-%m-%d').date() if return_date_str else None
        except ValueError:
            messages.error(request, "Invalid date format.")
            return render(request, 'assets/asset_edit.html', {
                'asset': asset,
                'branches': branches,
                'employees': employees,
                'cpu_choices': Asset.CPU_CHOICES,
                'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
                'ram_choices': Asset.RAM_CHOICES,
            })

        # Status rules
        if status == 'In Use' and not on_hand_date:
            messages.error(request, "On-Hand Date is required when status is 'In Use'.")
            return render(request, 'assets/asset_edit.html', {
                'asset': asset,
                'branches': branches,
                'employees': employees,
                'cpu_choices': Asset.CPU_CHOICES,
                'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
                'ram_choices': Asset.RAM_CHOICES,
            })

        if asset.status in ['In Use', 'Damage'] and status == 'Stock' and not return_date:
            messages.error(request, "Return Date is required when returning to Stock.")
            return render(request, 'assets/asset_edit.html', {
                'asset': asset,
                'branches': branches,
                'employees': employees,
                'cpu_choices': Asset.CPU_CHOICES,
                'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
                'ram_choices': Asset.RAM_CHOICES,
            })

        # Get related objects
        try:
            branch = Branch.objects.get(pk=branch_id)
        except Branch.DoesNotExist:
            messages.error(request, "Selected branch does not exist.")
            return render(request, 'assets/asset_edit.html', {
                'asset': asset,
                'branches': branches,
                'employees': employees,
                'cpu_choices': Asset.CPU_CHOICES,
                'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
                'ram_choices': Asset.RAM_CHOICES,
            })
        print(employee_id)
        employee = Employee.objects.get(pk=employee_id) if employee_id else None

        # Final update
        asset.product = product
        asset.serial = serial
        asset.status = status
        asset.type = asset_type
        asset.branch = branch
        asset.comments = comments
        asset.warranty = warranty
        asset.on_hand_date = on_hand_date if status == 'In Use' else None
        asset.return_date = return_date if status in ['Stock', 'Damage'] else None
        asset.employee_name = employee if status == 'In Use' else None
        asset.cpu = cpu
        asset.cpu_generation = cpu_gen
        asset.ram = ram
        asset._changed_by = request.user

        with transaction.atomic():
            asset.save()


        messages.success(request, "Asset updated successfully.")
        return redirect('edit_asset', asset_id=asset.pk)

    return render(request, 'assets/asset_edit.html', {
        'asset': asset,
        'branches': branches,
        'employees': employees,
        'cpu_choices': Asset.CPU_CHOICES,
        'cpu_gen_choices': Asset.CPU_GEN_CHOICES,
        'ram_choices': Asset.RAM_CHOICES,
    })







@login_required
def add_asset(request):
    if request.method == 'POST':
        form = AssetForm(request.POST)
        formset = StorageDeviceFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                asset = form.save(commit=False)
                asset.created_by = request.user
                asset.status = 'Stock'
                asset.branch = Branch.objects.get(slug='stock')
                asset._changed_by = request.user
                asset.save()

                # 🔴 KEY FIX: bind asset instance
                formset.instance = asset
                formset.save()

            messages.success(request, 'Asset added successfully.')
            return redirect('all_assets')

        else:
            messages.error(request, 'Please correct the errors below.')

    else:
        form = AssetForm()
        formset = StorageDeviceFormSet()

    return render(request, 'assets/asset_create.html', {
        'form': form,
        'formset': formset,
    })


##############################################################################    


'''
Create dynamic report view
This view allows users to select a model and apply filters to generate a report.
'''



# def dynamic_report_view(request):
#     report_model_id = request.GET.get("model")
#     filters = request.GET.copy()
#     filters.pop("model", None)

#     selected_model = None
#     fields = []
#     data = []

#     if report_model_id:
#         selected_model = ReportableModel.objects.get(id=report_model_id)
#         fields = ReportableField.objects.filter(model=selected_model, is_visible=True)

#         # Dynamically get model class
#         app_label, model_name = selected_model.model_path.split(".")
#         model_class = apps.get_model(app_label, model_name)

#         queryset = model_class.objects.all()

#         # Apply filters
#         for field in fields:
#             if field.is_filter:
#                 value = request.GET.get(field.field_name)
#                 if value:
#                     filter_key = field.field_name
#                     queryset = queryset.filter(**{filter_key: value})

#         data = queryset.values(*[f.field_name for f in fields])

#         # for obj in queryset:
#         #     row = {}
#         #     for field in fields:
#         #         field_name = field.field_name
#         #         value = getattr(obj, field_name)

#         #         # If it's a ForeignKey, get its string representation
#         #         if hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool)):
#         #             row[field_name] = str(value) if value else "None"
#         #         else:
#         #             row[field_name] = value

#         #     data.append(row)
#     context = {
#         "report_models": ReportableModel.objects.all(),
#         "selected_model": selected_model,
#         "fields": fields,
#         "data": data,
#         "filters": request.GET,
#     }

from django.apps import apps

#     return render(request, "reports/dynamic_report.html", context)
from django.shortcuts import render

from .models import ReportableField, ReportableModel  # Adjust import to your project


@login_required
def dynamic_report_view(request):
    report_model_id = request.GET.get("model")
    export_format = request.GET.get("export")
    filters = request.GET.copy()
    filters.pop("model", None)
    filters.pop("export", None)

    selected_model = None
    fields = []
    data = []

    if report_model_id:
        selected_model = ReportableModel.objects.get(id=report_model_id)
        fields = ReportableField.objects.filter(model=selected_model, is_visible=True)

        # Dynamically get model class
        app_label, model_name = selected_model.model_path.split(".")
        model_class = apps.get_model(app_label, model_name)

        queryset = model_class.objects.all()

        # Apply filters
        for field in fields:
            if field.is_filter:
                value = request.GET.get(field.field_name)
                if value:
                    filter_key = field.field_name
                    queryset = queryset.filter(**{filter_key: value})

        # Build data list with readable FK names
        for obj in queryset:
            row = {}
            for field in fields:
                field_name = field.field_name
                value = getattr(obj, field_name, None)

                if hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool)):
                    row[field_name] = str(value) if value else "None"
                else:
                    row[field_name] = value
            data.append(row)

        # Export to Excel if requested
        if export_format == 'excel':
            return export_report_to_excel(fields, data, selected_model.name)

    context = {
        "report_models": ReportableModel.objects.all(),
        "selected_model": selected_model,
        "fields": fields,
        "data": data,
        "filters": request.GET,
    }

    return render(request, "reports/dynamic_report.html", context)

def export_report_to_excel(fields, data, report_name="Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header Row
    header_font = Font(bold=True)
    for col_num, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num, value=field.display_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for row_num, row_data in enumerate(data, start=2):
        for col_num, field in enumerate(fields, 1):
            value = row_data.get(field.field_name, "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{report_name}.xlsx"'
    wb.save(response)
    return response



@login_required
def get_model_fields(request):
    model_path = request.GET.get('model_path')  # e.g. "assets.Asset"
    if not model_path or '.' not in model_path:
        return JsonResponse({'error': 'Invalid model path'}, status=400)

    try:
        app_label, model_name = model_path.split('.')
        model_class = apps.get_model(app_label, model_name)

        fields = []
        for field in model_class._meta.get_fields():
            if not field.auto_created and hasattr(field, 'name'):
                fields.append({
                    'name': field.name,
                    'verbose_name': getattr(field, 'verbose_name', field.name).title()
                })
        print(fields)
        return JsonResponse({'fields': fields})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

'''
screens & PC-screens part
'''

'''
Screens view
This view displays all screens and their statuses
'''
@login_required
def all_screens(request):
    screens = Screen.objects.select_related('employee', 'branch')
    print(screens)
    all_count = screens.count()
    screen_total = screens.filter(product='Screen').count()
    screen_pc_total = screens.filter(product='Screen-PC').count()
    unique_asset_types = list(
        screens.order_by('status').values_list('status', flat=True).distinct()
    )
    print(unique_asset_types)
    stock_count = screens.filter(status='Stock').count()
    in_use_count = screens.filter(status='In Use').count()
    damage_count = screens.filter(status='Damage').count()
    
    context = {
        'branch': None,
        'screens': screens,
        'all': all_count,
        'screen_total': screen_total,
        'screen_pc_total': screen_pc_total,
        "unique_asset_types": unique_asset_types,
        "stock_count": stock_count,
        "in_use_count": in_use_count,
        "damage_count": damage_count,
    }
    return render(request, 'screens/screens_all.html', context)

'''
Create screen view
This view allows users to create a new screen
'''
@login_required
def create_screen(request):
    if request.method == 'POST':
        form = ScreenForm(request.POST)
        if form.is_valid():
            screen = form.save(commit=False)
            screen.created_by = request.user
            screen._changed_by = request.user  # ✅ for signal
            screen.status = 'Stock'
            screen.branch = Branch.objects.get(slug='stock')
            screen.save()
            messages.success(request, f"Screen {screen.serial} created successfully.")
            return redirect('all_screens')
        else:
            messages.error(request, "Please correct the form errors.")
    else:
        form = ScreenForm()

    return render(request, 'screens/screens_create.html', {
        'form': form
    })


'''edit screen view
This view allows users to edit an existing screen'''

@login_required
def edit_screen(request, screen_id):
    screen = get_object_or_404(Screen, pk=screen_id)
    employees = Employee.objects.all()
    branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        product = request.POST.get('product')
        serial = request.POST.get('serial')
        brand = request.POST.get('brand')
        status = request.POST.get('status')
        employee_id = request.POST.get('employee')
        branch_id = request.POST.get('branch')

        employee = Employee.objects.get(pk=employee_id) if employee_id else None
        branch = Branch.objects.get(pk=branch_id) if branch_id else None

        # ✅ VALIDATION LOGIC
        if status == 'In Use':
            if not employee:
                messages.error(request, "Employee is required when status is 'In Use'.")
                return redirect('edit_screen', screen_id=screen.id)
            if branch and branch.name.lower() == 'stock':
                messages.error(request, "Branch must not be 'stock' when status is 'In Use'.")
                return redirect('edit_screen', screen_id=screen.id)

        elif status in ['Stock', 'Damage']:
            if employee:
                messages.error(request, "Employee must be empty when status is 'Stock' or 'Damage'.")
                return redirect('edit_screen', screen_id=screen.id)
            if not branch or branch.name.lower() != 'stock':
                messages.error(request, "Branch must be 'stock' when status is 'Stock' or 'Damage'.")
                return redirect('edit_screen', screen_id=screen.id)

        # ✅ Save changes
        screen.product = product
        screen.serial = serial
        screen.brand = brand
        screen.status = status
        screen.employee = employee if status == 'In Use' else None
        screen.branch = employee.branch if employee and status == 'In Use' else branch
        screen.updated_at = timezone.now()
        screen._changed_by = request.user
        screen.save()

        messages.success(request, f"Screen {serial} updated.")
        
        return redirect('edit_screen', screen_id=screen.id)

    return render(request, 'screens/screens_edit.html', {
        'screen': screen,
        'employees': employees,
        'branches': branches,
        'screen_product_choices': Screen.PRODUCT_CHOICES,
        'screen_status_choices': Screen.STATUS_CHOICES,
    })



''' unsign screen from employee
This view allows users to unassign a screen from an employee and return it to stock.'''
@require_POST
@login_required
def unassign_screen(request, screen_id, employee_id):
    screen = get_object_or_404(Screen, id=screen_id, employee__id=employee_id)
    stock_branch = Branch.objects.get(name__iexact='Stock')

    screen.employee = None
    screen.status = 'Stock'
    screen.branch = stock_branch
    screen.updated_at = timezone.now()
    screen.save()

    messages.success(request, f"Screen '{screen.serial}' unassigned successfully.")
    return redirect('edit_employee', employee_id=employee_id)

'''
Screen logs view
This view displays logs related to screen changes for a specific branch or all branches.
'''
@login_required
def all_screen_log(request, slug):
    if slug == 'All':
        logs = ScreenLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        screen_ids = Screen.objects.filter(branch=branch).values_list('id', flat=True)
        logs = ScreenLog.objects.filter(screen_id__in=screen_ids).order_by('-change_time')

    return render(request, 'screens/screens_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })
'''
branch screens view
This view displays all screens for a specific branch, including their statuses and types.
'''
@login_required
def branch_screens(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    screens = Screen.objects.filter(branch=branch)

    total = screens.count()
    screen_pcs = screens.filter(product='Screen-PC').count()
    regular_screens = screens.filter(product='Screen').count()

    return render(request, 'screens/branch_screens.html', {
        'screens': screens,
        'total': total,
        'screen_pcs': screen_pcs,
        'regular_screens': regular_screens,
        'current_branch_slug': slug,
    })

'''
Extract screens data
This view allows users to extract screen data based on selected status and format.
'''
@login_required
def extract_screens_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    
    return generate_screens_report(request, branch, selected_status, selected_format)

@login_required
def download_screens_template(request):
    return generate_screen_template()

@require_POST
@login_required
def upload_screens(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('all_screens')

    excel_file = request.FILES['excel_file']
    success = upload_bulk_screens(request, excel_file, branch, slug, request.user)
    if not success:
        messages.error(request, "Request failed, please try again")
        return redirect('all_screens')


    return redirect('all_screens')


def screen_details(request, screen_id):
    screen = get_object_or_404(Screen, pk=screen_id)
    print(screen)
    return render(request, 'screens/screens_details.html', {'screen': screen})

@require_POST
@login_required
def unassign_screen(request, screen_id, employee_id):
    screen = get_object_or_404(Screen, id=screen_id, employee__id=employee_id)
    stock_branch = Branch.objects.get(name__iexact='Stock')

    try:
        screen.employee = None
        screen.status = 'Stock'
        screen.branch = stock_branch
        screen.return_date = timezone.now().date()
        screen._changed_by = request.user
        screen.save()

        messages.success(request, f"Screen '{screen.serial}' unassigned successfully.")
    except Exception as e:
        messages.error(request, f"Error while unassigning screen: {e}")

    return redirect('edit_employee', employee_id=employee_id)

'''
Telephones part
'''

'''
Telephones view
This view displays all Telephones and their statuses
'''
@login_required
def all_telephones(request):
    telephones = Telephone.objects.select_related('employee', 'branch')
    print(telephones)
    all_count = telephones.count()
    telephone_total = telephones.count()
    unique_asset_types = list(
        telephones.order_by('status').values_list('status', flat=True).distinct()
    )
    print(unique_asset_types)
    stock_count = telephones.filter(status='Stock').count()
    in_use_count = telephones.filter(status='In Use').count()
    damage_count = telephones.filter(status='Damage').count()
    
    context = {
        'branch': None,
        'telephones': telephones,
        'all': all_count,
        'telephone_total': telephone_total,
        "unique_asset_types": unique_asset_types,
        "stock_count": stock_count,
        "in_use_count": in_use_count,
        "damage_count": damage_count,
    }
    return render(request, 'telephones/telephones_all.html', context)

'''
Create telephone view
This view allows users to create a new telephone
'''
@login_required
def create_telephone(request):
    if request.method == 'POST':
        form = TelephoneForm(request.POST)
        if form.is_valid():
            telephones = form.save(commit=False)
            telephones.created_by = request.user
            telephones._changed_by = request.user  # ✅ for signal
            telephones.status = 'Stock'
            telephones.branch = Branch.objects.get(slug='stock')
            telephones.save()
            messages.success(request, f"Telephone {telephones.serial} created successfully.")
            return redirect('all_telephones')
        else:
            messages.error(request, "Please correct the form errors.")
    else:
        form = TelephoneForm()

    return render(request, 'telephones/telephones_create.html', {
        'form': form
    })


'''edit telephone view
This view allows users to edit an existing telephone'''

@login_required
def edit_telephone(request, telephone_id):
    telephone = get_object_or_404(Telephone, pk=telephone_id)
    employees = Employee.objects.all()
    branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        
        product = request.POST.get('product')
        serial = request.POST.get('serial')
        brand = request.POST.get('brand')
        status = request.POST.get('status')
        employee_id = request.POST.get('employee')
        branch_id = request.POST.get('branch')

        employee = Employee.objects.filter(pk=employee_id).first() if employee_id else None
        branch = Branch.objects.filter(pk=branch_id).first() if branch_id else None

        # 🧠 Create a "preview" version of telephone with POST data
        telephone_preview = Telephone(
            id=telephone.id,  # preserve pk
            product=telephone.product,
            serial=telephone.serial,
            brand=brand,
            status=status,
            employee=employee,
            branch=branch or telephone.branch,
            created_at=telephone.created_at,
            updated_at=telephone.updated_at,
        )
        print(telephone_preview)
        context = {
            'telephone': telephone_preview,
            'employees': employees,
            'branches': branches,
            'telephone_status_choices': Telephone.STATUS_CHOICES,
        }

        # ✅ VALIDATION LOGIC
        if status == 'In Use':
            if not employee:
                messages.error(request, "Employee is required when status is 'In Use'.")
                return render(request, 'telephones/telephones_edit.html', context)
            if branch and branch.name.lower() == 'stock':
                messages.error(request, "Branch must not be 'stock' when status is 'In Use'.")
                return render(request, 'telephones/telephones_edit.html', context)

        elif status in ['Stock', 'Damage']:
            if employee:
                messages.error(request, "Employee must be empty when status is 'Stock' or 'Damage'.")
                return render(request, 'telephones/telephones_edit.html', context)
            if not branch or branch.name.lower() != 'stock':
                messages.error(request, "Branch must be 'stock' when status is 'Stock' or 'Damage'.")
                return render(request, 'telephones/telephones_edit.html', context)

        # ✅ Save to DB
        telephone.product = product
        telephone.serial = serial
        telephone.brand = brand
        telephone.status = status
        telephone.employee = employee if status == 'In Use' else None
        telephone.branch = employee.branch if employee and status == 'In Use' else branch
        telephone.updated_at = timezone.now()
        telephone._changed_by = request.user
        telephone.save()

        messages.success(request, f"Telephone {serial} updated.")
        return redirect('all_telephones')

    return render(request, 'telephones/telephones_edit.html', {
        'telephone': telephone,
        'employees': employees,
        'branches': branches,
        'telephone_status_choices': Telephone.STATUS_CHOICES,
    })



'''
Telephone logs view
This view displays logs related to telephone changes for a specific branch or all branches.
'''
@login_required
def all_telephone_log(request, slug):
    if slug == 'All':
        logs = TelephoneLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        telephone_ids = Screen.objects.filter(branch=branch).values_list('id', flat=True)
        logs = TelephoneLog.objects.filter(telephone_id__in=telephone_ids).order_by('-change_time')

    return render(request, 'telephones/telephones_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })
'''
branch telephones view
This view displays all telephones for a specific branch, including their statuses and types.
'''
@login_required
def branch_telephones(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    telephones = Telephone.objects.filter(branch=branch)

    total = telephones.count()


    return render(request, 'telephones/branch_telephones.html', {
        'telephones': telephones,
        'total': total,
        'current_branch_slug': slug,
    })

'''
Extract telephones data
This view allows users to extract screen data based on selected status and format.
'''
@login_required
def extract_telephones_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    
    return generate_telephones_report(request, branch, selected_status, selected_format)

@login_required
def download_telephones_template(request):
    return generate_telephone_template()

@require_POST
@login_required
def upload_telephones(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('all_telephones')

    excel_file = request.FILES['excel_file']
    success = upload_bulk_telephones(request, excel_file, branch, slug, request.user)
    if not success:
        messages.error(request, "Request failed, please try again")
        return redirect('all_telephones')


    return redirect('all_telephones')

@login_required
def telephone_details(request, telephone_id):
    telephone = get_object_or_404(Telephone, pk=telephone_id)
    print(telephone)
    return render(request, 'telephones/telephones_details.html', {'telephone': telephone})

''' unsign telephone from employee
This view allows users to unassign a telephone from an employee and return it to stock.'''
@require_POST
@login_required
def unassign_telephone(request, telephone_id, employee_id):
    telephone = get_object_or_404(Telephone, id=telephone_id, employee__id=employee_id)
    stock_branch = Branch.objects.get(name__iexact='Stock')

    telephone.employee = None
    telephone.status = 'Stock'
    telephone.branch = stock_branch
    telephone.updated_at = timezone.now()
    telephone.save()

    messages.success(request, f"Screen '{telephone.serial}' unassigned successfully.")
    return redirect('edit_employee', employee_id=employee_id)

'''Infra part'''
@login_required
def infrastructure_assets_view(request):

    return render(request, 'infra/infra.html')

'''Cameras part'''

@login_required
def all_cameras(request):
    unique_asset_types = list(
        Camera.objects.order_by('status').values_list('status', flat=True).distinct()
    )
    cameras = Camera.objects.all().order_by('id')
    all_count = cameras.count()
    stock_count = cameras.filter(status='Stock').count()
    in_use_count = cameras.filter(status='In Use').count()
    damage_count = cameras.filter(status='Damage').count()

    context = {
        'cameras': cameras,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        "unique_asset_types":unique_asset_types,
    }

    return render(request, 'infra/cameras/cameras_all.html', context)

@login_required
def branch_cameras(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    cameras = Camera.objects.filter(branch=branch)

    total = cameras.count()
    stock_count = cameras.filter(status='Stock').count()
    in_use_count = cameras.filter(status='In Use').count()
    damage_count = cameras.filter(status='Damage').count()

    context = {
        'branch': branch,
        'cameras': cameras,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }

    return render(request, 'infra/cameras/cameras_branch.html', context)

@xframe_options_exempt
def camera_details(request, camera_id):
    camera = get_object_or_404(Camera, pk=camera_id)
    print(camera)
    return render(request, 'infra/cameras/camera_details.html', {'camera': camera})

@login_required
def edit_camera(request, camera_id):
    camera = get_object_or_404(Camera, pk=camera_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = CameraEditForm(request.POST, instance=camera, user=request.user)

        if form.is_valid():
            camera = form.save(commit=False)
            status = form.cleaned_data.get('status')

            # Force branch to stock if not In Use
            if status != 'In Use':
                camera.branch = stock_branch
                camera.location='stock'
            else:
                # Ensure user selected a valid choosable branch
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    camera.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/cameras/camera_edit.html', {
                        'form': form,
                        'camera': camera,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                        'status': status,
                    })

            # Clear location if status changed from Stock to In Use
            if camera.status == 'Stock' and status == 'In Use':
                camera.location = None  # will be required via form validation

            camera._changed_by = request.user
            camera.save()
            
            return redirect('all_cameras')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        print(request.user)
        form = CameraEditForm(instance=camera, user=request.user)

    return render(request, 'infra/cameras/camera_edit.html', {
        'form': form,
        'camera': camera,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': camera.status,
    })

@login_required
def add_camera(request):
    if request.method == 'POST':
        form = CameraForm(request.POST)
        if form.is_valid():
            camera = form.save(commit=False)
            camera.created_by = request.user
            camera.status = 'Stock'
            camera.branch = Branch.objects.get(name__iexact='stock')
            camera._changed_by = request.user
            camera.save()

            messages.success(request, 'Camera added successfully.')
            return redirect('all_cameras')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CameraForm()

    return render(request, 'infra/cameras/camera_create.html', {
        'form': form,
    })
    
    
    
@login_required
def all_cameras_log(request, slug):
    if slug == 'All':
        logs = CameraLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        camera_ids = Camera.objects.filter(branch=branch).values_list('id', flat=True)
        logs = CameraLog.objects.filter(camera_id__in=camera_ids).order_by('-change_time')

    return render(request, 'infra/cameras/camera_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })
    
@require_POST 
@login_required
def upload_cameras(request, slug):
    print(slug)
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']

        success = upload_bulk_cameras(request, excel_file, branch, slug, request.user)

        if not success:
            if slug!='stock':
                return redirect('branch_assets', slug=slug)
            else:
                return redirect('all_cameras')

    else:
        messages.error(request, "No Excel file was uploaded.")

    if slug!='stock':
        return redirect('branch_assets', slug=slug)
    else:
        return redirect('all_cameras')

@login_required
def download_cameras_template(request):
    return generate_camera_template()

'''
Extract cameras data
This view allows users to extract screen data based on selected status and format.
'''
@login_required
def extract_cameras_data(request, slug):
    
    
    
    
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    
    return generate_cameras_report(request, branch, selected_status, selected_format)


''''
NVR part
'''
@login_required
def all_nvrs(request):
    unique_asset_types = list(
        NVR.objects.order_by('status').values_list('status', flat=True).distinct()
    )

    nvrs = NVR.objects.all().order_by('id')
    all_count = nvrs.count()
    stock_count = nvrs.filter(status='Stock').count()
    in_use_count = nvrs.filter(status='In Use').count()
    damage_count = nvrs.filter(status='Damage').count()

    context = {
        'nvrs': nvrs,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'unique_asset_types': unique_asset_types,
    }

    return render(request, 'infra/nvrs/nvr_all.html', context)


@login_required
def branch_nvrs(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    nvrs = NVR.objects.filter(branch=branch)

    total = nvrs.count()
    stock_count = nvrs.filter(status='Stock').count()
    in_use_count = nvrs.filter(status='In Use').count()
    damage_count = nvrs.filter(status='Damage').count()

    context = {
        'branch': branch,
        'nvrs': nvrs,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }

    return render(request, 'infra/nvrs/nvrs_by_branch.html', context)

@login_required
@xframe_options_exempt
def nvr_details(request, nvr_id):
    nvr = get_object_or_404(NVR, pk=nvr_id)
    return render(request, 'infra/nvrs/nvr_details.html', {'nvr': nvr})

@login_required
def edit_nvr(request, nvr_id):
    nvr = get_object_or_404(NVR, pk=nvr_id)

    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = NVREditForm(request.POST, instance=nvr)

        if form.is_valid():
            nvr = form.save(commit=False)
            status = form.cleaned_data.get('status')

            if status != 'In Use':
                nvr.branch = stock_branch
                nvr.location = 'stock'
            else:
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    nvr.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/nvrs/nvr_edit.html', {
                        'form': form,
                        'nvr': nvr,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                        'status': status,
                    })

            if nvr.status == 'Stock' and status == 'In Use':
                nvr.location = None

            nvr._changed_by = request.user
            nvr.save()
            return redirect('nvr_details', nvr_id=nvr.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = NVREditForm(instance=nvr)

    return render(request, 'infra/nvrs/nvr_edit.html', {
        'form': form,
        'nvr': nvr,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': nvr.status,
    })

@login_required
def add_nvr(request):
    if request.method == 'POST':
        form = NVRForm(request.POST)
        if form.is_valid():
            nvr = form.save(commit=False)
            nvr.created_by = request.user
            nvr.status = 'Stock'
            nvr.branch = Branch.objects.get(name__iexact='stock')
            nvr._changed_by = request.user
            nvr.save()

            messages.success(request, 'NVR added successfully.')
            return redirect('all_nvrs')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = NVRForm()

    return render(request, 'infra/nvrs/nvr_create.html', {'form': form})

@login_required
def all_nvr_logs(request, slug):
    if slug == 'All':
        logs = NVRLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        nvr_ids = NVR.objects.filter(branch=branch).values_list('id', flat=True)
        logs = NVRLog.objects.filter(nvr_id__in=nvr_ids).order_by('-change_time')

    return render(request, 'infra/nvrs/nvr_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })

@require_POST 
@login_required
def upload_nvrs(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        print(request.user)
        success = upload_bulk_nvrs(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('branch_assets', slug=slug) if slug != 'stock' else redirect('all_nvrs')
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('branch_assets', slug=slug) if slug != 'stock' else redirect('all_nvrs')

@login_required
def download_nvr_template(request):
    return generate_nvr_template()

@login_required
def extract_nvrs_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    
    return generate_nvrs_report(request, branch, selected_status, selected_format)

''' Firewall part'''

@login_required
def all_firewalls(request):
    unique_asset_types = list(
        Firewall.objects.order_by('status').values_list('status', flat=True).distinct()
    )
    firewalls = Firewall.objects.all().order_by('id')
    all_count = firewalls.count()
    stock_count = firewalls.filter(status='Stock').count()
    in_use_count = firewalls.filter(status='In Use').count()
    damage_count = firewalls.filter(status='Damage').count()

    context = {
        'firewalls': firewalls,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        "unique_asset_types": unique_asset_types,
    }

    return render(request, 'infra/firewalls/firewall_all.html', context)


@login_required
def branch_firewalls(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    firewalls = Firewall.objects.filter(branch=branch)

    context = {
        'branch': branch,
        'firewalls': firewalls,
        'all': firewalls.count(),
        'stock_count': firewalls.filter(status='Stock').count(),
        'in_use_count': firewalls.filter(status='In Use').count(),
        'damage_count': firewalls.filter(status='Damage').count(),
        'current_branch_slug': slug,
    }

    return render(request, 'infra/firewalls/firewall_branch.html', context)


@xframe_options_exempt
@login_required
def firewall_details(request, firewall_id):
    firewall = get_object_or_404(Firewall, pk=firewall_id)
    return render(request, 'infra/firewalls/firewall_details.html', {'firewall': firewall})


@login_required
def edit_firewall(request, firewall_id):
    firewall = get_object_or_404(Firewall, pk=firewall_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = FirewallEditForm(request.POST, instance=firewall)
        if form.is_valid():
            firewall = form.save(commit=False)
            status = form.cleaned_data.get('status')

            if status != 'In Use':
                firewall.branch = stock_branch
                firewall.location = 'stock'
            else:
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    firewall.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/firewalls/firewall_edit.html', {
                        'form': form,
                        'firewall': firewall,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                        'status': status,
                    })

            if firewall.status == 'Stock' and status == 'In Use':
                firewall.location = None

            firewall._changed_by = request.user
            firewall.save()

            return redirect('all_firewalls')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = FirewallEditForm(instance=firewall)

    return render(request, 'infra/firewalls/firewall_edit.html', {
        'form': form,
        'firewall': firewall,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': firewall.status,
    })


@login_required
def add_firewall(request):
    if request.method == 'POST':
        form = FirewallForm(request.POST)
        if form.is_valid():
            firewall = form.save(commit=False)
            firewall.created_by = request.user
            firewall.status = 'Stock'
            firewall.branch = Branch.objects.get(name__iexact='stock')
            firewall._changed_by = request.user
            firewall.save()
            messages.success(request, 'Firewall added successfully.')
            return redirect('all_firewalls')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FirewallForm()

    return render(request, 'infra/firewalls/firewall_create.html', {'form': form})


@login_required
def all_firewalls_log(request, slug):
    if slug == 'All':
        logs = FirewallLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        firewall_ids = Firewall.objects.filter(branch=branch).values_list('id', flat=True)
        logs = FirewallLog.objects.filter(firewall_id__in=firewall_ids).order_by('-change_time')

    return render(request, 'infra/firewalls/firewall_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })


@require_POST
@login_required
def upload_firewalls(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_firewalls(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('all_firewalls') if slug != 'stock' else redirect('all_firewalls')
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('branch_assets', slug=slug) if slug != 'stock' else redirect('all_firewalls')


@login_required
def download_firewalls_template(request):
    return generate_firewall_template()


@login_required
def extract_firewalls_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    return generate_firewalls_report(request, branch, selected_status, selected_format)

''''
Switch part
'''
#list all switchs
@login_required
def all_switches(request):
    switches = Switch.objects.all().order_by('id')
    all_count = switches.count()
    
    stock_count = switches.filter(status='Stock').count()
    in_use_count = switches.filter(status='In Use').count()
    damage_count = switches.filter(status='Damage').count()
    unique_switch_types = list(
        switches.order_by('status').values_list('status', flat=True).distinct()
    )
    return render(request, 'infra/switches/switch_all.html', {
        'switches': switches,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        "unique_switch_types":unique_switch_types,
    })

#list switch by branch
@login_required
def branch_switches(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    switches = Switch.objects.filter(branch=branch)

    return render(request, 'infra/switches/switches_branch.html', {
        'branch': branch,
        'switches': switches,
        'all': switches.count(),
        'stock_count': switches.filter(status='Stock').count(),
        'in_use_count': switches.filter(status='In Use').count(),
        'damage_count': switches.filter(status='Damage').count(),
        'current_branch_slug': slug,
    })

#switch details
@xframe_options_exempt
@login_required
def switch_details(request, switch_id):
    switch = get_object_or_404(Switch, pk=switch_id)
    return render(request, 'infra/switches/switch_details.html', {'switch': switch})

#add switch
@login_required
def create_switch(request):
    if request.method == 'POST':
        form = SwitchForm(request.POST)
        if form.is_valid():
            switch = form.save(commit=False)
            switch.created_by = request.user
            switch.status = 'Stock'
            switch.branch = Branch.objects.get(name__iexact='stock')
            switch.location = 'stock'
            switch._changed_by = request.user
            switch.save()

            messages.success(request, 'Switch added successfully.')
            return redirect('all_switches')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SwitchForm()

    return render(request, 'infra/switches/switch_create.html', {
        'form': form,
    })

#edit switch
@login_required
def edit_switch(request, switch_id):
    switch = get_object_or_404(Switch, pk=switch_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = SwitchEditForm(request.POST, instance=switch, user=request.user)
        if form.is_valid():
            switch = form.save(commit=False)
            status = form.cleaned_data.get('status')

            if status != 'In Use':
                switch.branch = stock_branch
                switch.location = 'stock'
            else:
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    switch.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/switches/switch_edit.html', {
                        'form': form,
                        'switch': switch,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                        'status': status,
                    })

            if switch.status == 'Stock' and status == 'In Use':
                switch.location = None

            switch._changed_by = request.user
            switch.save()
            return redirect('all_switches')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = SwitchEditForm(instance=switch, user=request.user)

    return render(request, 'infra/switches/switch_edit.html', {
        'form': form,
        'switch': switch,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': switch.status,
    })

#switch logs
@login_required
def all_switch_log(request, slug):
    if slug == 'All':
        logs = SwitchLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        switch_ids = Switch.objects.filter(branch=branch).values_list('id', flat=True)
        logs = SwitchLog.objects.filter(switch_id__in=switch_ids).order_by('-change_time')

    return render(request, 'infra/switches/switch_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })
    
#upload switches
@require_POST
@login_required
def upload_switches(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_switches(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('branch_switches', slug=slug) if slug != 'stock' else redirect('all_switches')
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('branch_switches', slug=slug) if slug != 'stock' else redirect('all_switches')

#download switch template
@login_required
def download_switches_template(request):
    return generate_switch_template()


#Extract switches data
@login_required
def extract_switches_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    return generate_switches_report(request, branch, selected_status, selected_format)

'''
Access Points part
'''
@login_required
def all_access_points(request):
    unique_asset_types = list(
        AccessPoint.objects.order_by('status').values_list('status', flat=True).distinct()
    )

    aps = AccessPoint.objects.all().order_by('id')
    all_count = aps.count()
    stock_count = aps.filter(status='Stock').count()
    in_use_count = aps.filter(status='In Use').count()
    damage_count = aps.filter(status='Damage').count()

    context = {
        'access_points': aps,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'unique_asset_types': unique_asset_types,
    }
    return render(request, 'infra/access_points/all_access_points.html', context)


@login_required
def branch_access_points(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    aps = AccessPoint.objects.filter(branch=branch)

    total = aps.count()
    stock_count = aps.filter(status='Stock').count()
    in_use_count = aps.filter(status='In Use').count()
    damage_count = aps.filter(status='Damage').count()

    context = {
        'branch': branch,
        'access_points': aps,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }
    return render(request, 'infra/access_points/access_points_by_branch.html', context)

@login_required
@xframe_options_exempt
def access_point_details(request, access_point_id):
    ap = get_object_or_404(AccessPoint, pk=access_point_id)
    return render(request, 'infra/access_points/access_point_details.html', {'access_point': ap})


@login_required
def edit_access_point(request, access_point_id):
    ap = get_object_or_404(AccessPoint, pk=access_point_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = AccessPointEditForm(request.POST, instance=ap, user=request.user)
        if form.is_valid():
            ap = form.save(commit=False)
            if ap.status != 'In Use':
                ap.branch = stock_branch
                ap.location = 'stock'
            else:
                if ap.branch and ap.branch.choosable:
                    pass
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/access_points/access_point_edit.html', {
                        'form': form,
                        'access_point': ap,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                    })

            ap._changed_by = request.user
            ap.save()
            return redirect('access_point_details', access_point_id=ap.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = AccessPointEditForm(instance=ap, user=request.user)

    return render(request, 'infra/access_points/access_point_edit.html', {
        'form': form,
        'access_point': ap,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
    })


@login_required
def create_access_point(request):
    if request.method == 'POST':
        form = AccessPointForm(request.POST)
        if form.is_valid():
            ap = form.save(commit=False)
            ap.created_by = request.user
            ap.status = 'Stock'
            ap.branch = Branch.objects.get(name__iexact='stock')
            ap._changed_by = request.user
            ap.save()
            messages.success(request, 'Access Point added successfully.')
            return redirect('all_access_points')
    else:
        form = AccessPointForm()

    return render(request, 'infra/access_points/access_point_create.html', {'form': form})


@login_required
def all_access_point_log(request, slug):
    if slug == 'All':
        logs = AccessPointLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        ap_ids = AccessPoint.objects.filter(branch=branch).values_list('id', flat=True)
        logs = AccessPointLog.objects.filter(access_point_id__in=ap_ids).order_by('-change_time')

    return render(request, 'infra/access_points/access_point_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })


@require_POST
@login_required
def upload_access_points(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_access_points(request, excel_file, branch, slug, request.user)
        if not success:
            return redirect('branch_assets', slug=slug)
    else:
        messages.error(request, "No Excel file was uploaded.")
    return redirect('branch_assets', slug=slug)


@login_required
def download_access_points_template(request):
    return generate_access_point_template()


@login_required
def extract_access_points_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')
    return generate_access_points_report(request, branch, selected_status, selected_format)

'''
Routers part
'''
@login_required
def all_routers(request):
    unique_asset_types = list(
        Router.objects.order_by('status').values_list('status', flat=True).distinct()
    )

    routers = Router.objects.all().order_by('id')
    all_count = routers.count()
    stock_count = routers.filter(status='Stock').count()
    in_use_count = routers.filter(status='In Use').count()
    damage_count = routers.filter(status='Damage').count()

    context = {
        'routers': routers,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'unique_asset_types': unique_asset_types,
    }

    return render(request, 'infra/routers/all_routers.html', context)


@login_required
def branch_routers(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    routers = Router.objects.filter(branch=branch)

    total = routers.count()
    stock_count = routers.filter(status='Stock').count()
    in_use_count = routers.filter(status='In Use').count()
    damage_count = routers.filter(status='Damage').count()

    context = {
        'branch': branch,
        'routers': routers,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }

    return render(request, 'infra/routers/routers_by_branch.html', context)


@xframe_options_exempt
@login_required
def router_details(request, router_id):
    router = get_object_or_404(Router, pk=router_id)
    return render(request, 'infra/routers/router_details.html', {'router': router})


@login_required
def edit_router(request, router_id):
    router = get_object_or_404(Router, pk=router_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = RouterEditForm(request.POST, instance=router, user=request.user)
        if form.is_valid():
            router = form.save(commit=False)
            status = form.cleaned_data.get('status')

            if status != 'In Use':
                router.branch = stock_branch
                router.location = 'stock'
            else:
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    router.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/routers/router_edit.html', {
                        'form': form,
                        'router': router,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                    })

            if router.status == 'Stock' and status == 'In Use':
                router.location = None

            router._changed_by = request.user
            router.save()
            return redirect('router_details', router_id=router.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = RouterEditForm(instance=router, user=request.user)

    return render(request, 'infra/routers/router_edit.html', {
        'form': form,
        'router': router,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': router.status,
    })


@login_required
def create_router(request):
    if request.method == 'POST':
        form = RouterForm(request.POST)
        if form.is_valid():
            router = form.save(commit=False)
            router.created_by = request.user
            router.status = 'Stock'
            router.branch = Branch.objects.get(name__iexact='stock')
            router._changed_by = request.user
            router.save()

            messages.success(request, 'Router added successfully.')
            return redirect('all_routers')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RouterForm()

    return render(request, 'infra/routers/router_create.html', {'form': form})


@login_required
def all_router_log(request, slug):
    if slug == 'All':
        logs = RouterLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        router_ids = Router.objects.filter(branch=branch).values_list('id', flat=True)
        logs = RouterLog.objects.filter(router_id__in=router_ids).order_by('-change_time')

    return render(request, 'infra/routers/router_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })


@require_POST
@login_required
def upload_routers(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_routers(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('branch_assets', slug=slug)
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('branch_assets', slug=slug)


@login_required
def download_routers_template(request):
    return generate_router_template()


@login_required
def extract_routers_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')

    return generate_routers_report(request, branch, selected_status, selected_format)

'''
UPS part
'''
@login_required
def all_ups(request):
    unique_ups_status = list(
        UPS.objects.order_by('status').values_list('status', flat=True).distinct()
    )

    ups_devices = UPS.objects.all().order_by('id')
    all_count = ups_devices.count()
    stock_count = ups_devices.filter(status='Stock').count()
    in_use_count = ups_devices.filter(status='In Use').count()
    damage_count = ups_devices.filter(status='Damage').count()

    context = {
        'ups_devices': ups_devices,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'unique_ups_status': unique_ups_status,
    }

    return render(request, 'infra/ups/ups_all.html', context)


@login_required
def branch_ups(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    ups_devices = UPS.objects.filter(branch=branch)

    total = ups_devices.count()
    stock_count = ups_devices.filter(status='Stock').count()
    in_use_count = ups_devices.filter(status='In Use').count()
    damage_count = ups_devices.filter(status='Damage').count()

    context = {
        'branch': branch,
        'ups_devices': ups_devices,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }

    return render(request, 'infra/ups/ups_by_branch.html', context)


@xframe_options_exempt
@login_required
def ups_details(request, ups_id):
    ups_device = get_object_or_404(UPS, pk=ups_id)
    return render(request, 'infra/ups/ups_details.html', {'ups_device': ups_device})


@login_required
def edit_ups(request, ups_id):
    ups_device = get_object_or_404(UPS, pk=ups_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = UPSEditForm(request.POST, instance=ups_device, user=request.user)
        if form.is_valid():
            ups_device = form.save(commit=False)
            status = form.cleaned_data.get('status')

            if status != 'In Use':
                ups_device.branch = stock_branch
                ups_device.location = 'stock'
            else:
                branch = form.cleaned_data.get('branch')
                if branch and branch.choosable:
                    ups_device.branch = branch
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/ups/ups_edit.html', {
                        'form': form,
                        'ups_device': ups_device,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                    })

            if ups_device.status == 'Stock' and status == 'In Use':
                ups_device.location = None

            ups_device._changed_by = request.user
            ups_device.save()
            return redirect('ups_details', ups_id=ups_device.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = UPSEditForm(instance=ups_device, user=request.user)

    return render(request, 'infra/ups/ups_edit.html', {
        'form': form,
        'ups_device': ups_device,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
        'status': ups_device.status,
    })


@login_required
def create_ups(request):
    if request.method == 'POST':
        form = UPSForm(request.POST)
        if form.is_valid():
            ups_device = form.save(commit=False)
            ups_device.created_by = request.user
            ups_device.status = 'Stock'
            ups_device.location = 'Stock'
            ups_device.branch = Branch.objects.get(name__iexact='stock')
            ups_device._changed_by = request.user
            ups_device.save()

            messages.success(request, 'UPS added successfully.')
            return redirect('all_ups')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UPSForm()

    return render(request, 'infra/ups/ups_create.html', {'form': form})


@login_required
def all_ups_log(request, slug):
    if slug == 'All':
        logs = UPSLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        ups_ids = UPS.objects.filter(branch=branch).values_list('id', flat=True)
        logs = UPSLog.objects.filter(ups_id__in=ups_ids).order_by('-change_time')

    return render(request, 'infra/ups/ups_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })


@require_POST
@login_required
def upload_ups(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_ups(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('all_ups', slug=slug)
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('all_ups', slug=slug)


@login_required
def download_ups_template(request):
    return generate_ups_template()


@login_required
def extract_ups_data(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')

    return generate_ups_report(request, branch, selected_status, selected_format)

'''
Raya data center part 
'''
@login_required
def all_raya_vms(request):
    vms = RayaDataCenterVM.objects.all().order_by('id')
    all_count = vms.count()
    uat_count = vms.filter(environment="uat").count()
    prod_count = vms.filter(environment="prod").count()
    expired_count = vms.filter(contract_end__lt=date.today()).count()

    context = {
        'vms': vms,
        'all': all_count,
        'uat_count': uat_count,
        'prod_count': prod_count,
        'expired': expired_count,
    }
    return render(request, 'raya/vm_all.html', context)



@xframe_options_exempt
@login_required
def raya_vm_details(request, vm_id):
    vm = get_object_or_404(RayaDataCenterVM, pk=vm_id)
    return render(request, 'raya/vm_details.html', {'vm': vm})


@login_required
def edit_raya_vm(request, vm_id):
    vm = get_object_or_404(RayaDataCenterVM, pk=vm_id)

    if request.method == 'POST':
        form = RayaDataCenterVMForm(request.POST, instance=vm, user=request.user)
        if form.is_valid():
            vm = form.save(commit=False)
            vm._changed_by = request.user
            vm.save()
            return redirect('vm_details', vm_id=vm.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = RayaDataCenterVMForm(instance=vm, user=request.user)

    return render(request, 'raya/vm_edit.html', {
        'form': form,
        'vm': vm,
        #'status': vm.status,
    })


@login_required
def create_raya_vm(request):
    if request.method == 'POST':
        form = RayaDataCenterVMForm(request.POST)
        if form.is_valid():
            vm = form.save(commit=False)
            vm.created_by = request.user
            vm.status = 'In Use'   # all Raya VMs are directly "In Use"
            vm._changed_by = request.user
            vm.save()

            messages.success(request, 'VM added successfully.')
            return redirect('all_vms')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RayaDataCenterVMForm()

    return render(request, 'raya/vm_create.html', {'form': form})




@require_POST
@login_required
def upload_raya_vms(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_raya_vms(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('all_vms')
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('all_vms')


@login_required
def download_raya_vm_template(request):
    return generate_raya_datacenter_template()


@login_required
def extract_vm_data(request):
    selected_format = request.POST.get('format')

    return generate_raya_report(request, selected_format)

"""
==========================
    ZK DEVICES VIEWS
==========================
"""

@login_required
def all_zk_devices(request):
    unique_statuses = list(
        ZKDevice.objects.order_by('status').values_list('status', flat=True).distinct()
    )
    devices = ZKDevice.objects.all().order_by('id')

    all_count = devices.count()
    stock_count = devices.filter(status='Stock').count()
    in_use_count = devices.filter(status='In Use').count()
    damage_count = devices.filter(status='Damage').count()

    context = {
        'devices': devices,
        'all': all_count,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        "unique_statuses": unique_statuses,
    }
    return render(request, 'infra/zk_devices/zk_all.html', context)


@login_required
def branch_zk_devices(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    devices = ZKDevice.objects.filter(branch=branch)

    total = devices.count()
    stock_count = devices.filter(status='Stock').count()
    in_use_count = devices.filter(status='In Use').count()
    damage_count = devices.filter(status='Damage').count()

    context = {
        'branch': branch,
        'devices': devices,
        'all': total,
        'stock_count': stock_count,
        'in_use_count': in_use_count,
        'damage_count': damage_count,
        'current_branch_slug': slug,
    }
    return render(request, 'infra/zk_devices/zk_branch.html', context)


@xframe_options_exempt
@login_required
def zk_device_details(request, device_id):
    device = get_object_or_404(ZKDevice, pk=device_id)
    return render(request, 'infra/zk_devices/zk_details.html', {'device': device})


@login_required
def edit_zk_device(request, device_id):
    zk = get_object_or_404(ZKDevice, pk=device_id)
    stock_branch = Branch.objects.filter(name__iexact='stock').first()
    choosable_branches = Branch.objects.filter(choosable=True)

    if request.method == 'POST':
        form = ZKDeviceEditForm(request.POST, instance=zk, user=request.user)
        if form.is_valid():
            zk = form.save(commit=False)
            if zk.status != 'In Use':
                zk.branch = stock_branch
                zk.location = 'stock'
            else:
                if zk.branch and zk.branch.choosable:
                    pass
                else:
                    messages.error(request, "Please select a valid branch.")
                    return render(request, 'infra/zk_devices/zk_edit.html', {
                        'form': form,
                        'device': zk,
                        'choosable_branches': choosable_branches,
                        'stock_branch': stock_branch,
                    })

            zk._changed_by = request.user
            zk.save()
            return redirect('zk_details', device_id=zk.id)
        else:
            print(form.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        form = ZKDeviceEditForm(instance=zk, user=request.user)

    return render(request, 'infra/zk_devices/zk_edit.html', {
        'form': form,
        'device': zk,
        'choosable_branches': choosable_branches,
        'stock_branch': stock_branch,
    })



@login_required
def add_zk_device(request):
    if request.method == 'POST':
        form = ZKDeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.created_by = request.user
            device.status = 'Stock'
            device.branch = Branch.objects.get(name__iexact='stock')
            device._changed_by = request.user
            device.save()

            messages.success(request, 'ZK Device added successfully.')
            return redirect('all_zk_devices')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ZKDeviceForm()

    return render(request, 'infra/zk_devices/zk_create.html', {'form': form})


@login_required
def zk_device_logs(request, slug):
    if slug == 'All':
        logs = ZKDeviceLog.objects.all().order_by('-change_time')
        branch = None
    else:
        branch = get_object_or_404(Branch, slug=slug)
        device_ids = ZKDevice.objects.filter(branch=branch).values_list('id', flat=True)
        logs = ZKDeviceLog.objects.filter(device_id__in=device_ids).order_by('-change_time')

    return render(request, 'infra/zk_devices/zk_logs.html', {
        'logs': logs,
        'branch': branch,
        'current_branch_slug': slug,
    })


@require_POST
@login_required
def upload_zk_devices(request, slug):
    branch = get_object_or_404(Branch, slug=slug)
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        success = upload_bulk_zk_devices(request, excel_file, branch, slug, request.user)

        if not success:
            return redirect('all_zk_devices', slug=slug) if slug != 'stock' else redirect('all_zk_devices')
    else:
        messages.error(request, "No Excel file was uploaded.")

    return redirect('all_zk_devices', slug=slug) if slug != 'stock' else redirect('all_zk_devices')


@login_required
def download_zk_template(request):
    return generate_zk_template()


@login_required
def extract_zk_devices(request, slug):
    branch = get_object_or_404(Branch, slug=slug) if slug != 'All' else 'All'
    selected_status = request.POST.get('status')
    selected_format = request.POST.get('format')

    return generate_zk_report(request, branch, selected_status, selected_format)

'''
Servers Part
'''
# List all servers
@login_required
def all_servers(request):
    servers = Server.objects.all().order_by("hostname")

    return render(request, "infra/servers/all_servers.html", {"servers": servers})

# Server details with nested VMs

@login_required
def server_details(request, server_id):
    server = get_object_or_404(Server, id=server_id)
    vms = server.vms.all().order_by("name")  # ✅ nested VMs
    return render(request, "infra/servers/server_details.html", {
        "server": server,
        "vms": vms
    })

# Add a new server
@login_required
def add_server(request):
    if request.method == "POST":
        form = ServerForm(request.POST)
        if form.is_valid():
            server = form.save(commit=False)
            server._changed_by = request.user  # Track the user
            server.save()
            messages.success(request, f"Server {server.hostname} added successfully.")
            return redirect("all_servers")
    else:
        form = ServerForm()
    return render(request, "infra/servers/server_create.html", {"form": form})
# views.py
@login_required
def server_resources(request, server_id):
    try:
        server = Server.objects.get(pk=server_id)
    except Server.DoesNotExist:
        raise Http404("Server not found")
    data = {
        "available_cpu": server.available_cpu_cores,
        "available_ram": server.available_ram_gb,
        "available_storage": server.available_storage_gb,
    }
    return JsonResponse(data)

# Add a new server
@login_required
def edit_server(request, server_id):
    server = get_object_or_404(Server, id=server_id)
    if request.method == "POST":
        form = ServerEditForm(request.POST, instance=server)  # ✅ use ServerEditForm
        if form.is_valid():
            server = form.save(commit=False)
            server._changed_by = request.user  # Track the user
            server.save()
            messages.success(request, f"Server {server.hostname} updated successfully.")
            return redirect("server_details", server_id=server.id)
        else:
            print(form.errors)
    else:
        form = ServerEditForm(instance=server)  # ✅
    return render(request, "infra/servers/edit_server.html", {"form": form, "server": server})



# Logs for a specific server

def server_logs(request):
    logs = ServerLog.objects.all().order_by('-change_time')  # Adjust model name & field
    return render(request, "infra/servers/server_logs.html", {"logs": logs})


'''
VMs part
'''
# List all VMs
@login_required
def all_vms(request):
    vms = VirtualMachine.objects.select_related("server").all().order_by("server__hostname", "name")
    return render(request, "infra/vms/all_vms.html", {"vms": vms})

# VM details with link back to parent server
@login_required
def vm_details(request, vm_id):
    vm = get_object_or_404(VirtualMachine, id=vm_id)
    return render(request, "infra/vms/vm_details.html", {"vm": vm, "server": vm.server})

# Add a new VM
@login_required
def add_vm(request):
    if request.method == "POST":
        form = VirtualMachineForm(request.POST)
        if form.is_valid():
            vm = form.save()
            messages.success(
                request, f"VM {vm.name} created successfully on {vm.server.hostname}."
            )
            return redirect("vm_details", vm_id=vm.id)
    else:
        form = VirtualMachineForm()
    return render(request, "infra/vms/vm_create.html", {"form": form})

# Edit an existing VM
@login_required
def edit_vm(request, vm_id):
    vm = get_object_or_404(VirtualMachine, id=vm_id)

    if request.method == "POST":
        form = VirtualMachineEditForm(request.POST, instance=vm, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"VM {vm.name} updated successfully.")
            return redirect("vm_details", vm_id=vm.id)
    else:
        form = VirtualMachineEditForm(instance=vm, user=request.user)

    return render(request, "infra/vms/edit_vm.html", {
        "form": form,
        "vm": vm
    })

# Logs for a specific VM
@login_required
def vm_logs(request, vm_id):
    vm = get_object_or_404(VirtualMachine, id=vm_id)
    logs = vm.logs.all()
    return render(request, "infra/vms/vm_logs.html", {"vm": vm, "logs": logs})
from django.views.generic import TemplateView


class NotificationDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = (
        'assets.view_notificationconfig',
        'assets.view_notificationrecipient',
    )
    template_name = 'notifications/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['configs'] = NotificationConfig.objects.all()
        context['recipients'] = NotificationRecipient.objects.all()
        return context

# class NotificationConfigListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
#     permission_required = 'assets.view_notificationconfig'
#     model = NotificationConfig
#     template_name = 'notifications/config_list.html'

class NotificationConfigCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'assets.add_notificationconfig'
    model = NotificationConfig
    form_class = NotificationConfigForm
    template_name = 'notifications/config_form.html'
    success_url = reverse_lazy('notification_dashboard')  # Redirect to dashboard
    
class NotificationConfigUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'assets.change_notificationconfig'
    model = NotificationConfig
    form_class = NotificationConfigForm
    template_name = 'notifications/config_form.html'
    success_url = reverse_lazy('notification_dashboard')  # Redirect to dashboard

class NotificationConfigDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'assets.delete_notificationconfig'
    model = NotificationConfig
    template_name = 'notifications/config_confirm_delete.html'
    success_url = reverse_lazy('notification_dashboard')  # Redirect to dashboard
    
# views.py



class RecipientCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'assets.add_notificationrecipient'
    model = NotificationRecipient
    form_class = NotificationRecipientForm
    template_name = 'notifications/recipient_form.html'
    success_url = reverse_lazy('notification_dashboard') 

class RecipientUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'assets.change_notificationrecipient'
    model = NotificationRecipient
    form_class = NotificationRecipientForm
    template_name = 'notifications/recipient_form.html'
    success_url = reverse_lazy('notification_dashboard')  

class RecipientDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'assets.delete_notificationrecipient'
    model = NotificationRecipient
    template_name = 'notifications/recipient_confirm_delete.html'
    success_url = reverse_lazy('notification_dashboard')  
