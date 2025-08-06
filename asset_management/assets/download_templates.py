from io import BytesIO

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (  # Make sure NVR is imported
    NVR,
    AccessPoint,
    Asset,
    Camera,
    Firewall,
    Router,
    Screen,
    StorageDevice,
    Switch,
    Telephone,
)

'''
Generate asset template
This function creates an Excel template for bulk asset upload with predefined headers and drop-downs.
'''
def generate_asset_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets Upload"

    # Define headers
    headers = [
        'Product', 'Serial', 'Type', 'CPU', 'CPU Generation', 'RAM',
        'Warranty', 'Comments',
        'Storage 1 Type', 'Storage 1 Size',
        'Storage 2 Type', 'Storage 2 Size'
    ]
    ws.append(headers)

    # Highlight mandatory fields
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col in ['A', 'B', 'C']:  # Product, Serial, Type
        ws[f"{col}1"].fill = red_fill

    # Define drop-downs (choices from model)
    def add_dropdown(col_letter, choices):
        dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}100")

    add_dropdown("C", [c[0] for c in Asset.ASSET_TYPE_CHOICES])     # Type
    add_dropdown("D", [c[0] for c in Asset.CPU_CHOICES])            # CPU
    add_dropdown("E", [c[0] for c in Asset.CPU_GEN_CHOICES])        # CPU Gen
    add_dropdown("F", [c[0] for c in Asset.RAM_CHOICES])            # RAM
    add_dropdown("I", [c[0] for c in StorageDevice.STORAGE_TYPE_CHOICES])  # Storage 1 Type
    add_dropdown("J", [c[0] for c in StorageDevice.STORAGE_SIZE_CHOICES])  # Storage 1 Size
    add_dropdown("K", [c[0] for c in StorageDevice.STORAGE_TYPE_CHOICES])  # Storage 2 Type
    add_dropdown("L", [c[0] for c in StorageDevice.STORAGE_SIZE_CHOICES])  # Storage 2 Size

    # Return file as HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asset_template.xlsx'
    return response

'''
Generate screen template
This function creates an Excel template for bulk screen upload with predefined headers and drop-down
'''
def generate_screen_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screens Upload"

    headers = ['Product', 'Serial','Brand']
    ws.append(headers)

    # Highlight required fields
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col in ['A', 'B','C']:
        ws[f"{col}1"].fill = red_fill

    # Add dropdown for Product
    def add_dropdown(col_letter, choices):
        dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}100")

    add_dropdown("A", [c[0] for c in Screen.PRODUCT_CHOICES])

    # Return file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=screen_template.xlsx'
    return response

'''
Generate telephone template
This function creates an Excel template for bulk telephone upload with predefined headers and drop-down
'''
def generate_telephone_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Telephones Upload"

    headers = ['Product', 'Serial','Brand']
    ws.append(headers)

    # Highlight required fields
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col in ['A', 'B','C']:
        ws[f"{col}1"].fill = red_fill

    # Return file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=telephone_template.xlsx'
    return response

''' Generate camera template
This function creates an Excel template for bulk NVRs upload with predefined headers'''
def generate_camera_template():
    wb = openpyxl.Workbook()

    # ==================== CAMERA SHEET ====================
    ws_cam = wb.active
    ws_cam.title = "Camera Upload"
    camera_headers = [
        'Model', 'Serial Number', 'Power Source', 
        'IP Address', 'MAC Address','Purchase Date'
    ]
    ws_cam.append(camera_headers)

    # Required: Model, Serial Number, Power Source
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B', 'C']:
        ws_cam[f"{col_letter}1"].fill = red_fill

    # Dropdown for Status
    camera_status_choices = [c[0] for c in Camera.STATUS_CHOICES]
    dv_status_cam = DataValidation(type="list", formula1=f'"{",".join(camera_status_choices)}"', allow_blank=False)
    ws_cam.add_data_validation(dv_status_cam)
    dv_status_cam.add("G2:G100")


    # ==================== RETURN FILE ====================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_camera_template.xlsx'
    return response



''' Generate NVR template
This function creates an Excel template for bulk NVRs upload with predefined headers'''
def generate_nvr_template():
    wb = openpyxl.Workbook()

    # ==================== NVR SHEET ====================
    ws_nvr = wb.active
    ws_nvr.title = "NVR Upload"
    nvr_headers = [
        'Model', 'Serial Number', 'HDD Capacity', 'Number of Ports',
        'IP Address', 'MAC Address', 'Purchase Date'
    ]
    ws_nvr.append(nvr_headers)

    # Required: Model, Serial Number, HDD Capacity, Number of Ports
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B', 'C', 'D']:
        ws_nvr[f"{col_letter}1"].fill = red_fill

    # Optional: add dropdown for status if needed
    nvr_status_choices = [choice[0] for choice in NVR.STATUS_CHOICES]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(nvr_status_choices)}"', allow_blank=False)
    ws_nvr.add_data_validation(dv_status)
    dv_status.add("H2:H100")  # Column H = Status (if included)

    # ==================== RETURN FILE ====================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_nvr_template.xlsx'
    return response


''' Generate Firewall template
This function creates an Excel template for bulk firewalls upload with predefined headers'''

def generate_firewall_template():
    wb = openpyxl.Workbook()

    # ========== FIREWALL SHEET ==========
    ws_fw = wb.active
    ws_fw.title = "Firewall Upload"
    fw_headers = [
        'Model', 'Serial Number', 'Firmware Version', 'Number of Ports',
        'IP Address', 'MAC Address', 'License Expiry Date', 'Purchase Date'
    ]
    ws_fw.append(fw_headers)

    # Mark required fields in red: Model, Serial Number
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B']:
        ws_fw[f"{col_letter}1"].fill = red_fill

    # Optional: Dropdown for status (if needed in the future)
    fw_status_choices = [choice[0] for choice in Firewall.STATUS_CHOICES]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(fw_status_choices)}"', allow_blank=False)
    ws_fw.add_data_validation(dv_status)
    dv_status.add("I2:I100")  # I column = Status (not used here, but for future-proofing)

    # ========== RETURN FILE ==========
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_firewall_template.xlsx'
    return response

'''
Generate switch template
This function creates an Excel template for bulk switch upload with predefined headers and drop-downs.
'''
def generate_switch_template():
    wb = openpyxl.Workbook()

    # ==================== SWITCH SHEET ====================
    ws_switch = wb.active
    ws_switch.title = "Switch Upload"

    switch_headers = [
        'Model', 'Serial Number', 'Number of Ports', 'Number of POE Ports',
        'IP Address', 'MAC Address', 'Purchase Date'
    ]
    ws_switch.append(switch_headers)

    # Highlight required fields: Model, Serial Number, Number of Ports, Number of POE Ports
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B', 'C', 'D']:
        ws_switch[f"{col_letter}1"].fill = red_fill

    # Integer-only for Number of Ports (C) and POE Ports (D)
    dv_integer_ports = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
    ws_switch.add_data_validation(dv_integer_ports)
    dv_integer_ports.add("C2:C100")

    dv_integer_poe = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
    ws_switch.add_data_validation(dv_integer_poe)
    dv_integer_poe.add("D2:D100")

    # ========== RETURN FILE ==========
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_switch_template.xlsx'
    return response

'''
Generate Access Point template  
This function creates an Excel template for bulk access point upload with predefined headers
'''
def generate_access_point_template():
    wb = openpyxl.Workbook()

    # ========== ACCESS POINT SHEET ==========
    ws_ap = wb.active
    ws_ap.title = "Access Point Upload"
    ap_headers = [
        'Model', 'Serial Number', 'IP Address', 'MAC Address', 'Expiry Date', 'Purchase Date'
    ]
    ws_ap.append(ap_headers)

    # Mark required fields in red: Model, Serial Number
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B']:
        ws_ap[f"{col_letter}1"].fill = red_fill

    # Optional: Dropdown for status (future-proof)
    ap_status_choices = [choice[0] for choice in AccessPoint.STATUS_CHOICES]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(ap_status_choices)}"', allow_blank=False)
    ws_ap.add_data_validation(dv_status)
    dv_status.add("G2:G100")  # Placeholder if status column added later

    # ========== RETURN FILE ==========
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_access_point_template.xlsx'
    return response

'''
Generate Router template  
This function creates an Excel template for bulk router upload with predefined headers
'''
def generate_router_template():
    wb = openpyxl.Workbook()

    # ========== ROUTER SHEET ==========
    ws_r = wb.active
    ws_r.title = "Router Upload"
    router_headers = [
        'Model', 'Serial Number', 'IP Address', 'MAC Address', 'Purchase Date'
    ]
    ws_r.append(router_headers)

    # Mark required fields in red: Model, Serial Number
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col_letter in ['A', 'B']:
        ws_r[f"{col_letter}1"].fill = red_fill

    # Optional: Dropdown for status (future-proof)
    router_status_choices = [choice[0] for choice in Router.STATUS_CHOICES]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(router_status_choices)}"', allow_blank=False)
    ws_r.add_data_validation(dv_status)
    dv_status.add("F2:F100")  # Placeholder if status column added later

    # ========== RETURN FILE ==========
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=infra_router_template.xlsx'
    return response
