"""Excel import / export / template plumbing shared by the asset endpoints.

Each asset type keeps its own column definitions (see ``EXCEL_COLUMNS`` in the
per-type API modules); only the mechanical spreadsheet handling lives here.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

from django.apps import apps
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from .openapi import (
    EXCEL_IMPORT_RESPONSE,
    EXCEL_UPLOAD_REQUEST,
    XLSX_FILE_RESPONSE,
)

MAX_REPORTED_ERRORS = 100
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

HEADER_FILL = PatternFill("solid", start_color="1F3B73", end_color="1F3B73")
REQUIRED_FILL = PatternFill("solid", start_color="B3341F", end_color="B3341F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass(frozen=True)
class Column:
    """One spreadsheet column bound to a serializer field.

    ``lookup`` maps a human value (branch name, employee email) onto a related
    row, so spreadsheets never have to carry database ids.
    """

    header: str
    field: str
    source: str = ""
    lookup: tuple[str, str] | None = None
    required: bool = False

    @property
    def export_path(self) -> str:
        return self.source or self.field


def _clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _normalise(value):
    """Turn a spreadsheet cell into something a serializer accepts."""
    value = _clean(value)
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _resolve_path(instance, path):
    for part in path.split("."):
        if instance is None:
            return None
        instance = getattr(instance, part, None)
        if callable(instance):
            instance = instance()
    if isinstance(instance, (dt.datetime, dt.date)):
        return instance.replace(tzinfo=None) if isinstance(instance, dt.datetime) else instance
    return instance


def _workbook_response(workbook, filename):
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _style_header(worksheet, columns):
    for index, column in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=index, value=column.header)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if column.required else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            14, min(len(column.header) + 6, 40)
        )
    worksheet.freeze_panes = "A2"


class ExcelIOMixin:
    """Adds ``/excel-template/``, ``/export/`` and ``/import/`` to a ViewSet."""

    excel_columns: list[Column] = []
    excel_sheet_name = "Data"

    def _field_choices(self, field_name):
        field = self.get_serializer().fields.get(field_name)
        if isinstance(field, serializers.ChoiceField):
            return [str(value) for value in field.choices]
        return []

    def _lookup_values(self, column, limit=200):
        if not column.lookup:
            return []
        model_path, attribute = column.lookup
        model = apps.get_model(model_path)
        return [
            str(value)
            for value in model.objects.values_list(attribute, flat=True).order_by(
                attribute
            )[:limit]
            if value
        ]

    @extend_schema(
        summary="Download an import template",
        description=(
            "Returns an empty .xlsx workbook with the expected column headers. "
            "Required columns are highlighted and choice columns carry "
            "drop-down validation."
        ),
        responses={200: XLSX_FILE_RESPONSE},
    )
    @action(detail=False, methods=["get"], url_path="excel-template")
    def excel_template(self, request, *args, **kwargs):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.excel_sheet_name[:31]
        _style_header(worksheet, self.excel_columns)

        for index, column in enumerate(self.excel_columns, start=1):
            options = self._field_choices(column.field) or self._lookup_values(column)
            if not options:
                continue
            # Excel caps inline list validations, so only short lists are inlined.
            joined = ",".join(options)
            if len(joined) > 250:
                continue
            validation = DataValidation(
                type="list", formula1=f'"{joined}"', allow_blank=not column.required
            )
            worksheet.add_data_validation(validation)
            letter = get_column_letter(index)
            validation.add(f"{letter}2:{letter}1000")

        filename = f"{self.basename}_template.xlsx"
        return _workbook_response(workbook, filename)

    @extend_schema(
        summary="Export to Excel",
        description=(
            "Exports the current result set as .xlsx. All list filters, "
            "`search` and `ordering` query parameters apply, so the export "
            "matches exactly what the list endpoint returns."
        ),
        responses={200: XLSX_FILE_RESPONSE},
    )
    @action(detail=False, methods=["get"])
    def export(self, request, *args, **kwargs):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.excel_sheet_name[:31]
        _style_header(worksheet, self.excel_columns)

        queryset = self.filter_queryset(self.get_queryset())
        for row_number, instance in enumerate(queryset.iterator(), start=2):
            for index, column in enumerate(self.excel_columns, start=1):
                worksheet.cell(
                    row=row_number,
                    column=index,
                    value=_resolve_path(instance, column.export_path),
                )

        stamp = dt.date.today().isoformat()
        return _workbook_response(workbook, f"{self.basename}_{stamp}.xlsx")

    @extend_schema(
        summary="Bulk import from Excel",
        description=(
            "Uploads a workbook produced from the import template. Related "
            "rows are matched by natural key (branch name, employee email, "
            "server hostname) so spreadsheets never carry database ids.\n\n"
            "Valid rows are saved and invalid rows are reported individually, "
            "so a single bad row does not block the rest of the file."
        ),
        request=EXCEL_UPLOAD_REQUEST,
        responses={200: EXCEL_IMPORT_RESPONSE, 400: EXCEL_IMPORT_RESPONSE},
    )
    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request, *args, **kwargs):
        upload = request.FILES.get("file")
        if upload is None:
            return Response(
                {"detail": "Attach an .xlsx file in the 'file' field."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not upload.name.lower().endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = load_workbook(upload, data_only=True, read_only=True)
        except Exception as exc:
            return Response(
                {"detail": f"Could not read the workbook: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return Response(
                {"detail": "The workbook is empty."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = {
            str(value).strip().lower(): index
            for index, value in enumerate(header_row)
            if value is not None
        }
        columns = [c for c in self.excel_columns if c.header.lower() in headers]
        missing_required = [
            c.header
            for c in self.excel_columns
            if c.required and c.header.lower() not in headers
        ]
        if missing_required:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing_required)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        errors = []
        for row_number, row in enumerate(rows, start=2):
            if row is None or all(_clean(cell) is None for cell in row):
                continue

            payload = {}
            row_errors = {}
            for column in columns:
                raw = row[headers[column.header.lower()]]
                value = _normalise(raw)
                if value is None:
                    continue
                if column.lookup:
                    model_path, attribute = column.lookup
                    model = apps.get_model(model_path)
                    match = model.objects.filter(
                        **{f"{attribute}__iexact": str(value)}
                    ).first()
                    if match is None:
                        row_errors[column.field] = [
                            f"No {model._meta.verbose_name} with {attribute} '{value}'."
                        ]
                        continue
                    value = match.pk
                payload[column.field] = value

            if not row_errors:
                serializer = self.get_serializer(data=payload)
                if serializer.is_valid():
                    serializer.save()
                    created += 1
                    continue
                row_errors = serializer.errors

            if len(errors) < MAX_REPORTED_ERRORS:
                errors.append({"row": row_number, "errors": row_errors})

        return Response(
            {
                "created": created,
                "failed": len(errors),
                "errors": errors,
            },
            status=status.HTTP_200_OK if created else status.HTTP_400_BAD_REQUEST,
        )
