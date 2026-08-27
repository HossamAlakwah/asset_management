"""Smoke tests for permissions, logs, search/filter, and Excel I/O."""

from io import BytesIO

import pyotp
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from assets.models import (
    Branch,
    Camera,
    CameraLog,
    Employee,
    Laptop,
    LaptopLog,
    NotificationConfig,
)
from users.twofactor import confirm_setup, start_setup

User = get_user_model()


@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
class SmokeTest(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="smoke.admin",
            password="smoke-pass-123",
            role="super_admin",
            is_staff=True,
            is_superuser=True,
        )
        self.viewer = User.objects.create_user(
            username="smoke.user",
            password="smoke-pass-123",
            role="user",
        )
        self.branch = Branch.objects.create(name="Smoke HQ")
        self.other_branch = Branch.objects.create(name="Smoke Downtown")
        self.employee = Employee.objects.create(
            name="Smoke Employee",
            department="IT",
            title="Engineer",
            email="smoke.employee@example.com",
            branch=self.branch,
            created_by=self.admin,
        )
        self.client = APIClient()

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def test_anonymous_is_blocked(self):
        self.assertEqual(self.client.get("/api/v1/laptops/").status_code, 401)
        self.assertEqual(self.client.get("/api/v1/schema/").status_code, 401)
        self.assertEqual(self.client.get("/app/").status_code, 302)

    def test_viewer_can_read_but_not_write(self):
        self.auth(self.viewer)
        self.assertEqual(self.client.get("/api/v1/laptops/").status_code, 200)
        self.assertEqual(self.client.get("/api/v1/schema/").status_code, 200)
        create = self.client.post(
            "/api/v1/laptops/",
            {"product": "ThinkPad", "serial": "SMOKE-DENIED", "status": "Stock"},
            format="json",
        )
        self.assertEqual(create.status_code, 403)
        users = self.client.get("/api/v1/users/")
        self.assertEqual(users.status_code, 403)

    def test_admin_can_write_and_super_admin_sees_users(self):
        self.auth(self.admin)
        created = self.client.post(
            "/api/v1/laptops/",
            {"product": "ThinkPad T14", "serial": "SMOKE-LT-ADMIN", "status": "Stock"},
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        self.assertEqual(self.client.get("/api/v1/users/").status_code, 200)

    def test_laptop_logs_create_assign_unassign(self):
        self.auth(self.admin)
        created = self.client.post(
            "/api/v1/laptops/",
            {
                "product": "ThinkPad E14",
                "serial": "SMOKE-LT-LOG",
                "status": "Stock",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        laptop_id = created.data["id"]
        logs = LaptopLog.objects.filter(laptop_id=laptop_id)
        self.assertEqual(logs.count(), 1)
        self.assertEqual(logs.first().new_status, "Stock")
        self.assertEqual(logs.first().changed_by_id, self.admin.id)

        assigned = self.client.patch(
            f"/api/v1/laptops/{laptop_id}/",
            {"employee": self.employee.id},
            format="json",
        )
        self.assertEqual(assigned.status_code, 200, assigned.data)
        laptop = Laptop.objects.get(pk=laptop_id)
        self.assertEqual(laptop.status, "In Use")
        self.assertEqual(laptop.branch_id, self.branch.id)
        self.assertIsNotNone(laptop.on_hand_date)
        self.assertEqual(LaptopLog.objects.filter(laptop_id=laptop_id).count(), 2)
        assign_log = LaptopLog.objects.filter(laptop_id=laptop_id).order_by("id").last()
        self.assertEqual(assign_log.old_status, "Stock")
        self.assertEqual(assign_log.new_status, "In Use")
        self.assertEqual(assign_log.new_employee_id, self.employee.id)

        unassigned = self.client.post(f"/api/v1/laptops/{laptop_id}/unassign/")
        self.assertEqual(unassigned.status_code, 200, unassigned.data)
        laptop.refresh_from_db()
        self.assertEqual(laptop.status, "Stock")
        self.assertIsNone(laptop.employee_id)
        self.assertEqual(LaptopLog.objects.filter(laptop_id=laptop_id).count(), 3)
        last = LaptopLog.objects.filter(laptop_id=laptop_id).order_by("id").last()
        self.assertEqual(last.new_status, "Stock")
        self.assertIsNone(last.new_employee_id)

        retrieve = self.client.get(f"/api/v1/laptops/{laptop_id}/")
        self.assertEqual(retrieve.status_code, 200)
        self.assertEqual(len(retrieve.data["logs"]), 3)

    def test_spec_change_is_not_a_lifecycle_log(self):
        """Hardware edits currently do not write a history row."""
        self.auth(self.admin)
        laptop = Laptop.objects.create(
            product="Old name", serial="SMOKE-LT-SPEC", status="Stock", branch=self.branch
        )
        before = LaptopLog.objects.filter(laptop=laptop).count()
        self.client.patch(
            f"/api/v1/laptops/{laptop.id}/",
            {"product": "New name", "ram": "16GB"},
            format="json",
        )
        self.assertEqual(LaptopLog.objects.filter(laptop=laptop).count(), before)

    def test_camera_logs_location_and_status(self):
        self.auth(self.admin)
        created = self.client.post(
            "/api/v1/cameras/",
            {
                "model": "Hikvision",
                "serial_number": "SMOKE-CAM-1",
                "status": "Stock",
                "power_source": "PoE",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        camera_id = created.data["id"]
        self.assertEqual(CameraLog.objects.filter(camera_id=camera_id).count(), 1)

        moved = self.client.patch(
            f"/api/v1/cameras/{camera_id}/",
            {"status": "In Use", "location": "Lobby"},
            format="json",
        )
        self.assertEqual(moved.status_code, 200, moved.data)
        logs = list(CameraLog.objects.filter(camera_id=camera_id).order_by("id"))
        self.assertEqual(len(logs), 2)
        self.assertEqual(logs[-1].old_status, "Stock")
        self.assertEqual(logs[-1].new_status, "In Use")
        self.assertEqual(logs[-1].new_location, "Lobby")

    def test_search_and_filter(self):
        self.auth(self.admin)
        Laptop.objects.create(
            product="ThinkPad T14",
            serial="SMOKE-SRCH-T14",
            status="Stock",
            branch=self.branch,
        )
        Laptop.objects.create(
            product="Latitude 5540",
            serial="SMOKE-SRCH-LAT",
            status="In Use",
            employee=self.employee,
            branch=self.branch,
        )
        search = self.client.get("/api/v1/laptops/?search=SMOKE-SRCH-T14")
        self.assertEqual(search.status_code, 200)
        serials = [row["serial"] for row in search.data["results"]]
        self.assertEqual(serials, ["SMOKE-SRCH-T14"])

        filtered = self.client.get("/api/v1/laptops/?status=Stock&search=SMOKE-SRCH")
        self.assertEqual(filtered.status_code, 200)
        serials = [row["serial"] for row in filtered.data["results"]]
        self.assertEqual(serials, ["SMOKE-SRCH-T14"])

        by_branch = self.client.get(f"/api/v1/laptops/?branch={self.other_branch.id}")
        self.assertEqual(by_branch.status_code, 200)
        self.assertEqual(by_branch.data["count"], 0)

    def test_excel_template_export_import(self):
        self.auth(self.admin)
        Laptop.objects.create(
            product="Export Me",
            serial="SMOKE-XLS-OUT",
            status="Stock",
            branch=self.branch,
        )
        template = self.client.get("/api/v1/laptops/excel-template/")
        self.assertEqual(template.status_code, 200)
        self.assertIn(
            "spreadsheetml.sheet",
            template["Content-Type"],
        )

        export = self.client.get("/api/v1/laptops/export/?search=SMOKE-XLS-OUT")
        self.assertEqual(export.status_code, 200)
        exported = load_workbook(BytesIO(export.content))
        rows = list(exported.active.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 2)
        self.assertIn("Serial", rows[0])
        self.assertTrue(any(row and "SMOKE-XLS-OUT" in row for row in rows[1:]))

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Product", "Serial", "Status", "Branch"])
        sheet.append(["Imported Box", "SMOKE-XLS-IN", "Stock", self.branch.name])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        payload.name = "laptops.xlsx"

        viewer = APIClient()
        viewer.force_authenticate(user=self.viewer)
        denied = viewer.post("/api/v1/laptops/import/", {"file": payload}, format="multipart")
        self.assertEqual(denied.status_code, 403)

        payload.seek(0)
        imported = self.client.post(
            "/api/v1/laptops/import/", {"file": payload}, format="multipart"
        )
        self.assertEqual(imported.status_code, 200, imported.data)
        self.assertEqual(imported.data["created"], 1)
        self.assertEqual(imported.data["failed"], 0)
        laptop = Laptop.objects.get(serial="SMOKE-XLS-IN")
        self.assertEqual(laptop.product, "Imported Box")
        self.assertEqual(laptop.branch_id, self.branch.id)
        self.assertTrue(LaptopLog.objects.filter(laptop=laptop).exists())

        missing_assignee = Workbook()
        missing_sheet = missing_assignee.active
        missing_sheet.append(["Product", "Serial", "Status", "Branch"])
        missing_sheet.append(["Used Box", "SMOKE-XLS-USE", "In Use", self.branch.name])
        missing_payload = BytesIO()
        missing_assignee.save(missing_payload)
        missing_payload.seek(0)
        missing_payload.name = "laptops-in-use.xlsx"
        missing_import = self.client.post(
            "/api/v1/laptops/import/", {"file": missing_payload}, format="multipart"
        )
        self.assertEqual(missing_import.status_code, 400, missing_import.data)
        self.assertEqual(missing_import.data["created"], 0)
        self.assertFalse(Laptop.objects.filter(serial="SMOKE-XLS-USE").exists())

        assigned_book = Workbook()
        assigned_sheet = assigned_book.active
        assigned_sheet.append(["Product", "Serial", "Status", "Branch", "Employee Email"])
        assigned_sheet.append(
            [
                "Assigned Box",
                "SMOKE-XLS-ASN",
                "In Use",
                self.branch.name,
                self.employee.email,
            ]
        )
        assigned_payload = BytesIO()
        assigned_book.save(assigned_payload)
        assigned_payload.seek(0)
        assigned_payload.name = "laptops-assigned.xlsx"
        assigned_import = self.client.post(
            "/api/v1/laptops/import/", {"file": assigned_payload}, format="multipart"
        )
        self.assertEqual(assigned_import.status_code, 200, assigned_import.data)
        assigned = Laptop.objects.get(serial="SMOKE-XLS-ASN")
        self.assertEqual(assigned.status, "In Use")
        self.assertEqual(assigned.employee_id, self.employee.id)
        logs = list(LaptopLog.objects.filter(laptop=assigned).order_by("id"))
        self.assertEqual([row.new_status for row in logs], ["Stock", "In Use"])

        short_book = Workbook()
        short_sheet = short_book.active
        short_sheet.append(["Product", "Serial", "Status", "Branch", "Employee Email"])
        short_sheet.append(["Short Box", "SMOKE-XLS-SHORT"])
        short_payload = BytesIO()
        short_book.save(short_payload)
        short_payload.seek(0)
        short_payload.name = "laptops-short.xlsx"
        short_import = self.client.post(
            "/api/v1/laptops/import/", {"file": short_payload}, format="multipart"
        )
        self.assertEqual(short_import.status_code, 200, short_import.data)
        self.assertEqual(short_import.data["created"], 1)
        self.assertEqual(short_import.data["failed"], 0)
        self.assertTrue(Laptop.objects.filter(serial="SMOKE-XLS-SHORT").exists())

    def test_create_status_validation_and_in_use_logs(self):
        self.auth(self.admin)
        denied_in_use = self.client.post(
            "/api/v1/laptops/",
            {
                "product": "Forced Stock",
                "serial": "SMOKE-LT-FORCE",
                "status": "In Use",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(denied_in_use.status_code, 400, denied_in_use.data)
        self.assertIn("employee", denied_in_use.data)

        denied_damage = self.client.post(
            "/api/v1/screens/",
            {
                "product": "Screen",
                "serial": "SMOKE-SCR-FORCE",
                "brand": "Dell",
                "status": "Damage",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(denied_damage.status_code, 400, denied_damage.data)
        self.assertIn("status", denied_damage.data)

        denied_camera = self.client.post(
            "/api/v1/cameras/",
            {
                "model": "Forced Cam",
                "serial_number": "SMOKE-CAM-FORCE",
                "status": "In Use",
                "power_source": "PoE",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(denied_camera.status_code, 400, denied_camera.data)
        self.assertIn("status", denied_camera.data)

        created = self.client.post(
            "/api/v1/laptops/",
            {
                "product": "Assigned on create",
                "serial": "SMOKE-LT-ASN",
                "status": "In Use",
                "employee": self.employee.id,
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        self.assertEqual(created.data["status"], "In Use")
        self.assertEqual(created.data["employee"], self.employee.id)
        logs = list(
            LaptopLog.objects.filter(laptop_id=created.data["id"]).order_by("id")
        )
        self.assertEqual([row.new_status for row in logs], ["Stock", "In Use"])

        schema = self.client.get("/api/v1/schema/")
        laptops = next(item for item in schema.data["resources"] if item["key"] == "laptops")
        status_field = next(field for field in laptops["fields"] if field["name"] == "status")
        self.assertTrue(status_field["omit_on_create"])
        vms = next(
            item for item in schema.data["resources"] if item["key"] == "virtual-machines"
        )
        vm_status = next(field for field in vms["fields"] if field["name"] == "status")
        self.assertFalse(vm_status.get("omit_on_create"))

    def test_ported_original_validations(self):
        self.auth(self.admin)
        camera = self.client.post(
            "/api/v1/cameras/",
            {
                "model": "Hikvision",
                "serial_number": "SMOKE-CAM-LOC",
                "power_source": "PoE",
                "branch": self.branch.id,
            },
            format="json",
        )
        self.assertEqual(camera.status_code, 201, camera.data)
        denied_location = self.client.patch(
            f"/api/v1/cameras/{camera.data['id']}/",
            {"status": "In Use"},
            format="json",
        )
        self.assertEqual(denied_location.status_code, 400, denied_location.data)
        self.assertIn("location", denied_location.data)

        placed = self.client.patch(
            f"/api/v1/cameras/{camera.data['id']}/",
            {"status": "In Use", "location": "Lobby"},
            format="json",
        )
        self.assertEqual(placed.status_code, 200, placed.data)

        contract = self.client.post(
            "/api/v1/colocation-vms/",
            {
                "name": "smoke-dc-bad-dates",
                "ip_address": "10.20.30.40",
                "vcpu": 2,
                "vram_gb": 4,
                "environment": "prod",
                "contract_start": "2026-06-01",
                "contract_end": "2026-01-01",
            },
            format="json",
        )
        self.assertEqual(contract.status_code, 400, contract.data)
        self.assertIn("contract_end", contract.data)

        server = self.client.post(
            "/api/v1/servers/",
            {
                "model": "Dell R740",
                "serial_number": "SMOKE-SRV-1",
                "hostname": "smoke-host-1",
                "cpu_cores": 8,
                "ram_gb": 32,
                "storage_gb": 500,
                "hypervisor": "vmware",
            },
            format="json",
        )
        self.assertEqual(server.status_code, 400, server.data)
        self.assertIn("ip_address", server.data)

        server = self.client.post(
            "/api/v1/servers/",
            {
                "model": "Dell R740",
                "serial_number": "SMOKE-SRV-1",
                "hostname": "smoke-host-1",
                "ip_address": "10.10.10.10",
                "cpu_cores": 8,
                "ram_gb": 32,
                "storage_gb": 500,
                "hypervisor": "vmware",
            },
            format="json",
        )
        self.assertEqual(server.status_code, 201, server.data)

        vm = self.client.post(
            "/api/v1/virtual-machines/",
            {
                "server": server.data["id"],
                "name": "smoke-vm-1",
                "vcpu": 1,
                "vram_gb": 2,
                "storage_gb": 20,
            },
            format="json",
        )
        self.assertEqual(vm.status_code, 400, vm.data)
        self.assertIn("ip_address", vm.data)

    def test_schema_hides_split_alert_pages_and_exposes_alerts(self):
        self.auth(self.admin)
        schema = self.client.get("/api/v1/schema/")
        self.assertEqual(schema.status_code, 200)
        keys = {item["key"] for item in schema.data["resources"]}
        self.assertNotIn("notification-configs", keys)
        self.assertNotIn("notification-recipients", keys)
        self.assertNotIn("raya-vms", keys)
        self.assertIn("colocation-vms", keys)
        titles = {item["key"]: item["title"] for item in schema.data["resources"]}
        self.assertEqual(titles["colocation-vms"], "Colocation VMs")
        self.assertIn("alerts", schema.data)
        self.assertTrue(schema.data["alerts"]["models"])

    def test_notification_create_with_recipients_and_run(self):
        self.auth(self.admin)
        Laptop.objects.create(
            product="Stock box", serial="SMOKE-STK-1", status="Stock", branch=self.branch
        )
        created = self.client.post(
            "/api/v1/notification-configs/",
            {
                "model_name": "Laptop",
                "condition_type": "stock_count",
                "condition_value": "99",
                "is_active": True,
                "notification_message": "Low {model}: {count}/{threshold}",
                "recipients": ["alerts@example.com"],
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        self.assertEqual(created.data["recipients"], ["alerts@example.com"])
        run = self.client.post(f"/api/v1/notification-configs/{created.data['id']}/run/")
        self.assertEqual(run.status_code, 200, run.data)
        self.assertTrue(run.data["triggered"])
        self.assertEqual(run.data["emailed"], 1)

        rejected = self.client.post(
            "/api/v1/notification-configs/",
            {
                "model_name": "Desktop",
                "condition_type": "stock_count",
                "condition_value": "not-a-number",
                "is_active": True,
            },
            format="json",
        )
        self.assertEqual(rejected.status_code, 400, rejected.data)
        self.assertIn("condition_value", rejected.data)

        bad = NotificationConfig.objects.create(
            model_name="Laptop",
            condition_type="stock_count",
            condition_value="not-a-number",
            is_active=True,
        )
        bad_run = self.client.post(f"/api/v1/notification-configs/{bad.id}/run/")
        self.assertEqual(bad_run.status_code, 200, bad_run.data)
        self.assertFalse(bad_run.data["triggered"])
        self.assertEqual(bad_run.data["emailed"], 0)
        self.assertIn("whole number", bad_run.data["reason"])

    def test_pages_load(self):
        login = self.client.get("/login/")
        self.assertEqual(login.status_code, 200)
        docs = self.client.get("/api/docs/")
        self.assertEqual(docs.status_code, 200)
        device = start_setup(self.admin)
        confirm_setup(self.admin, pyotp.TOTP(device.secret).now())
        self.assertTrue(
            self.client.login(username="smoke.admin", password="smoke-pass-123")
        )
        session = self.client.session
        session["two_factor_verified"] = True
        session.save()
        app = self.client.get("/app/")
        self.assertEqual(app.status_code, 200)
        stats = self.client.get("/api/v1/stats/")
        self.assertEqual(stats.status_code, 200)
        self.assertIn("cards", stats.data)
        self.assertIn("alerts", stats.data)
